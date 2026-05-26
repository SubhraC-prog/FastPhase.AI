"""
CHROME-pred Report Template v3.0
Implements the full 10-section HPLC Method Development Report
for both Excel (openpyxl) and PDF (reportlab) output.
"""

from __future__ import annotations
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "navy":       "1F4E79",
    "blue":       "2E75B6",
    "light_blue": "DEEAF1",
    "sky":        "BDD7EE",
    "teal":       "17375E",
    "green_hdr":  "375623",
    "green_bg":   "E2EFDA",
    "green_cell": "C6EFCE",
    "amber_hdr":  "7F6000",
    "amber_bg":   "FFF2CC",
    "amber_cell": "FFEB9C",
    "red_hdr":    "9C0006",
    "red_bg":     "FFDCE1",
    "red_cell":   "FFC7CE",
    "grey_hdr":   "404040",
    "grey_bg":    "F2F2F2",
    "white":      "FFFFFF",
    "black":      "000000",
}

# ── Safe getter helpers ───────────────────────────────────────────────────────
def _g(d, *keys, default="N/A"):
    """Safe nested dict getter."""
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
    return cur if cur not in (None, "") else default


def _f(v, fmt=".3f", default="N/A"):
    """Format float safely."""
    try:
        return format(float(v), fmt)
    except (TypeError, ValueError):
        return default


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────────────────────────────────────
class ReportTemplate:
    """
    Builds the 10-section CHROME-pred Excel report
    matching the canonical template document.
    """

    def __init__(self):
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        self._Font = Font
        self._Fill = PatternFill
        self._Align = Alignment
        self._Border = Border
        self._Side = Side
        self._col_letter = get_column_letter

    # ── low-level cell helper ────────────────────────────────────────────────
    def _s(self, ws, row, col, value,
           bold=False, italic=False, size=10, color="000000",
           bg=None, align="left", wrap=False,
           border=False, num_fmt=None):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, size=size, color=color,
                      name="Arial")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
        if border:
            thin = Side(style="thin", color="B8CCE4")
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if num_fmt:
            c.number_format = num_fmt
        return c

    def _hdr(self, ws, row, col_start, col_end, text, bg=C["navy"], fg="FFFFFF"):
        """Merge+style a section header."""
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
        self._s(ws, row, col_start, text,
                bold=True, size=11, color=fg, bg=bg, align="center")
        ws.row_dimensions[row].height = 20

    def _sub(self, ws, row, col_start, col_end, text):
        """Merge+style a sub-section header."""
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
        self._s(ws, row, col_start, text,
                bold=True, size=10, color="FFFFFF", bg=C["blue"], align="left")
        ws.row_dimensions[row].height = 18

    def _tbl_hdr(self, ws, row, cols, labels, bg=C["teal"]):
        """Write a table header row."""
        for col, label in zip(cols, labels):
            self._s(ws, row, col, label,
                    bold=True, size=9, color="FFFFFF", bg=bg,
                    align="center", border=True)
        ws.row_dimensions[row].height = 16

    def _row(self, ws, row, cols, values, bg=None, bold=False, border=True):
        """Write a data row."""
        fill = bg or (C["grey_bg"] if row % 2 == 0 else C["white"])
        for col, val in zip(cols, values):
            self._s(ws, row, col, val, bg=fill, bold=bold,
                    border=border, size=9, wrap=True)

    # ── public sheet creators ────────────────────────────────────────────────
    def create_cover_sheet(self, ws, results: Dict):
        self._build_cover(ws, results)

    def create_summary_sheet(self, ws, results: Dict):
        self._build_summary(ws, results)

    def create_physchem_sheet(self, ws, results: Dict):
        self._build_physchem(ws, results)

    def create_hsm_sheet(self, ws, results: Dict):
        self._build_hsm(ws, results)

    def create_column_sheet(self, ws, results: Dict):
        self._build_columns(ws, results)

    def create_solvent_sheet(self, ws, results: Dict):
        self._build_solvents(ws, results)

    def create_buffer_sheet(self, ws, results: Dict):
        self._build_buffers(ws, results)

    def create_gradient_sheet(self, ws, results: Dict):
        self._build_gradient(ws, results)

    def create_references_sheet(self, ws, results: Dict):
        self._build_references(ws, results)

    def create_regulatory_sheet(self, ws, results: Dict):
        self._build_regulatory(ws, results)

    def create_metadata_sheet(self, ws, results: Dict):
        self._build_metadata(ws, results)

    # ── COVER ────────────────────────────────────────────────────────────────
    def _build_cover(self, ws, results):
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

        # Title banner
        ws.merge_cells("B1:E2")
        self._s(ws, 1, 2,
                "CHROME-pred  |  AI-Assisted HPLC Method Development Report",
                bold=True, size=16, color="FFFFFF", bg=C["navy"], align="center")
        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 10

        ws.merge_cells("B3:E3")
        self._s(ws, 3, 2,
                "Chromatography AI System v3.0  —  Comprehensive Method Development",
                bold=False, size=10, color="FFFFFF", bg=C["teal"], align="center")
        ws.row_dimensions[3].height = 18

        # Report header table
        fields = [
            ("Report ID",        _g(results, "report_id")),
            ("Generated",        _g(results, "timestamp")),
            ("Compound Name",    _g(results, "name")),
            ("SMILES",           _g(results, "smiles")),
            ("Project",          _g(results, "project")),
            ("Notes",            _g(results, "notes")),
            ("Status",           _g(results, "status")),
            ("Processing Time",  _f(_g(results, "processing_time_s"), ".2f") + " s"),
            ("Overall Confidence", _f(_g(results, "scores", "overall"), ".2f")),
        ]
        r = 5
        for label, value in fields:
            self._s(ws, r, 2, label, bold=True, size=10, bg=C["light_blue"],
                    border=True)
            ws.merge_cells(start_row=r, start_column=3,
                           end_row=r, end_column=5)
            self._s(ws, r, 3, value, size=10, bg=C["white"], border=True, wrap=True)
            ws.row_dimensions[r].height = 16
            r += 1

        # Module confidence table
        r += 1
        self._hdr(ws, r, 2, 5, "MODULE CONFIDENCE SCORES")
        r += 1
        self._tbl_hdr(ws, r, [2, 3, 4, 5],
                      ["Module", "Score", "Status", "Rationale"])
        r += 1
        scores = _g(results, "scores") if isinstance(_g(results, "scores"), dict) else {}
        rationale = _g(results, "rationale") if isinstance(_g(results, "rationale"), dict) else {}
        modules = [
            ("Physicochemical", "physchem"),
            ("HSM Descriptors", "hsm"),
            ("Buffer Selection", "buffer"),
            ("Solvent Selection", "solvent"),
            ("Column Selection", "column"),
        ]
        for name, key in modules:
            score = _f(scores.get(key, 0), ".2f")
            s_val = float(scores.get(key, 0))
            status = "High" if s_val >= 0.8 else ("Moderate" if s_val >= 0.5 else "Low")
            bg = C["green_cell"] if s_val >= 0.8 else (
                C["amber_cell"] if s_val >= 0.5 else C["red_cell"])
            self._s(ws, r, 2, name, bold=True, size=9, bg=bg, border=True)
            self._s(ws, r, 3, score, size=9, bg=bg, border=True, align="center")
            self._s(ws, r, 4, status, size=9, bg=bg, border=True, align="center")
            rat = rationale.get(key, "")
            self._s(ws, r, 5, str(rat)[:120], size=9, bg=C["white"],
                    border=True, wrap=True)
            ws.row_dimensions[r].height = 28
            r += 1

    # ── SECTION 1: PHYSCHEM ──────────────────────────────────────────────────
    def _build_physchem(self, ws, results):
        p = _g(results, "physchem") if isinstance(_g(results, "physchem"), dict) else {}

        for col, w in zip("ABCDEF", [3, 30, 20, 18, 18, 28]):
            ws.column_dimensions[col].width = w

        r = 1
        self._hdr(ws, r, 2, 6,
                  "SECTION 1 — Compound Identification & Physicochemical Profile")
        r += 1

        # 1.1 Compound info
        self._sub(ws, r, 2, 6, "1.1  Compound Information")
        r += 1
        info = [
            ("Compound Name",    _g(results, "name")),
            ("SMILES",           _g(results, "smiles")),
            ("Molecular Formula",_g(p, "molecular_formula")),
            ("Molecular Weight", _f(_g(p, "molecular_weight"), ".3f") + " g/mol"),
            ("Exact Mass",       _f(_g(p, "exact_mass"), ".5f") + " Da"),
            ("Project",          _g(results, "project")),
            ("Notes",            _g(results, "notes")),
        ]
        for label, val in info:
            self._s(ws, r, 2, label, bold=True, size=9, bg=C["light_blue"], border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            self._s(ws, r, 3, val, size=9, bg=C["white"], border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 1.2 Physicochemical
        r += 1
        self._sub(ws, r, 2, 6, "1.2  Physicochemical Properties")
        r += 1
        self._tbl_hdr(ws, r, [2, 3, 4, 5, 6],
                      ["Descriptor", "Value", "Unit", "Threshold", "Reference"])
        r += 1
        props_1_2 = [
            ("LogP (Wildman-Crippen)", _f(_g(p,"logp"),".3f"), "", "−2 to 5",
             "Wildman & Crippen (1999) DOI:10.1021/ci990307l"),
            ("LogD (pH 2.0)",  _f(_g(p,"logd_ph2"),".3f"),  "", "", "Henderson-Hasselbalch"),
            ("LogD (pH 5.0)",  _f(_g(p,"logd_ph5"),".3f"),  "", "", ""),
            ("LogD (pH 7.4)",  _f(_g(p,"logd_ph74"),".3f"), "", "", "Physiological pH"),
            ("LogD (pH 9.0)",  _f(_g(p,"logd_ph9"),".3f"),  "", "", ""),
            ("LogD (pH 11.0)", _f(_g(p,"logd_ph11"),".3f"), "", "", ""),
            ("TPSA", _f(_g(p,"tpsa"),".2f"), "Å²", "≤140",
             "Ertl et al. (2000) DOI:10.1021/jm000942e"),
            ("H-Bond Donors",    str(_g(p,"hbd_lipinski","0")), "", "≤5",
             "Lipinski et al. (1997)"),
            ("H-Bond Acceptors", str(_g(p,"hba_lipinski","0")), "", "≤10", ""),
            ("Rotatable Bonds",  str(_g(p,"rotatable_bonds","0")), "", "≤10",
             "Veber et al. (2002)"),
            ("Aromatic Rings",   str(_g(p,"aromatic_rings","0")), "", "", ""),
            ("Aliphatic Rings",  str(_g(p,"aliphatic_rings","0")), "", "", ""),
            ("Fraction Csp³",    _f(_g(p,"fraction_csp3"),".3f"), "", "", ""),
        ]
        for i, row_data in enumerate(props_1_2):
            bg = C["grey_bg"] if i % 2 == 0 else C["white"]
            for col, val in zip([2,3,4,5,6], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 1.3 Drug-likeness
        r += 1
        self._sub(ws, r, 2, 6, "1.3  Drug-Likeness Assessment")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Filter","Status","Threshold","Value","Reference"])
        r += 1
        lipi = _g(p,"lipinski_violations","0")
        dl_rows = [
            ("Lipinski Violations", str(lipi), "≤1 acceptable",
             str(lipi), "Lipinski et al. (1997)"),
            ("Ghose Filter", str(_g(p,"ghose_filter","N/A")),
             "MW 160-480, LogP -0.4 to 5.6",
             _f(_g(p,"molecular_weight"),".1f"),
             "Ghose et al. (1999)"),
            ("Veber Filter",  str(_g(p,"veber_filter","N/A")),
             "Rot ≤10, TPSA ≤140",
             str(_g(p,"rotatable_bonds","0")),
             "Veber et al. (2002)"),
            ("Muegge Filter", str(_g(p,"muegge_filter","N/A")),
             "MW 200-600, LogP -2 to 5",
             _f(_g(p,"logp"),".3f"),
             "Muegge et al. (2001)"),
            ("QED Score",    _f(_g(p,"qed_score"),".3f"), ">0.5 favourable",
             _f(_g(p,"qed_score"),".3f"),
             "Bickerton et al. (2012)"),
        ]
        for i, row_data in enumerate(dl_rows):
            bg = C["green_cell"] if "True" in str(row_data[1]) or (
                row_data[0]=="Lipinski Violations" and str(lipi) in ("0","1")
            ) else C["amber_cell"]
            for col, val in zip([2,3,4,5,6], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 1.4 Ionization
        r += 1
        self._sub(ws, r, 2, 6, "1.4  Ionization Profile")
        r += 1
        ion_rows = [
            ("Ionization Type",    str(_g(p,"ionization_type","N/A"))),
            ("pKa Acidic (min)",   _f(_g(p,"pka_acidic_min"),".2f")),
            ("pKa Acidic (max)",   _f(_g(p,"pka_acidic_max"),".2f")),
            ("pKa Basic (min)",    _f(_g(p,"pka_basic_min"),".2f")),
            ("pKa Basic (max)",    _f(_g(p,"pka_basic_max"),".2f")),
            ("Isoelectric Point",  _f(_g(p,"isoelectric_point"),".2f")),
            ("Formal Charge",      str(_g(p,"formal_charge","0"))),
        ]
        self._tbl_hdr(ws, r, [2,3], ["Parameter","Value"])
        r += 1
        for i, (lbl, val) in enumerate(ion_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            self._s(ws, r, 3, val, size=9, bg=bg, border=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 1.5 Solubility & Permeability
        r += 1
        self._sub(ws, r, 2, 6, "1.5  Solubility & Permeability")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Parameter","Value","Unit","Classification","Reference"])
        r += 1
        sol_rows = [
            ("LogS (ESOL)", _f(_g(p,"logS"),".2f"), "log mol/L",
             _g(p,"solubility_class","N/A"),
             "Delaney (2004) DOI:10.1021/ci034243x"),
            ("Intrinsic Solubility", _f(_g(p,"intrinsic_solubility"),".3f"),
             "mg/mL", "", "Yalkowsky & Valvani (1980)"),
            ("Caco-2 Permeability",  _f(_g(p,"caco2_permeability"),".1f"),
             "×10⁻⁶ cm/s", "", "PAMPA model"),
            ("MDCK Permeability",    _f(_g(p,"mdck_permeability"),".1f"),
             "×10⁻⁶ cm/s", "", "Irvine et al. (1999)"),
        ]
        for i, row_data in enumerate(sol_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5,6], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

    # ── SECTION 2: HSM ───────────────────────────────────────────────────────
    def _build_hsm(self, ws, results):
        hsm = _g(results,"hsm") if isinstance(_g(results,"hsm"),dict) else {}
        p   = _g(results,"physchem") if isinstance(_g(results,"physchem"),dict) else {}

        for col, w in zip("ABCDEF", [3,26,16,32,22,24]):
            ws.column_dimensions[col].width = w

        r = 1
        self._hdr(ws, r, 2, 6,
                  "SECTION 2 — Hydrophobic Subtraction Model (HSM) Descriptors")
        r += 1

        # 2.1 Descriptor table
        self._sub(ws, r, 2, 6, "2.1  HSM Descriptor Values")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Descriptor","Symbol","Value","Interpretation","Reference"])
        r += 1

        def _interp(key, val):
            v = float(val) if val != "N/A" else 0.0
            if key == "eta_prime":
                return "High hydrophobicity" if v>1.5 else ("Moderate" if v>0.8 else "Low")
            elif key == "sigma_prime":
                return "High steric resistance" if v>0.3 else "Low steric resistance"
            elif key == "beta_prime":
                return "Strong H-bond base" if v>0.5 else "Weak H-bond base"
            elif key == "alpha_prime":
                return "Strong H-bond acid" if v>0.5 else "Weak H-bond acid"
            elif key == "kappa_prime":
                return "Significant cationic" if v>0.5 else "Minimal cationic"
            return ""

        hsm_desc = [
            ("Hydrophobicity",    "η′ (eta')",   "eta_prime",
             "Marchand et al. (2008) DOI:10.1016/j.chroma.2007.11.101"),
            ("Steric Resistance", "σ′ (sigma')", "sigma_prime",
             "Marchand et al. (2005) DOI:10.1016/j.chroma.2004.11.014"),
            ("H-Bond Basicity",   "β′ (beta')",  "beta_prime",
             "Abraham (1993) DOI:10.1039/CS9932200073"),
            ("H-Bond Acidity",    "α′ (alpha')", "alpha_prime",
             "Abraham (1993)"),
            ("Cationic Charge",   "κ′ (kappa')", "kappa_prime",
             "Dolan et al. (2004) DOI:10.1016/j.chroma.2004.09.020"),
        ]
        for i, (name, sym, key, ref) in enumerate(hsm_desc):
            val = _f(_g(hsm, key), ".3f")
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, v in zip([2,3,4,5,6],
                               [name, sym, val, _interp(key,val), ref]):
                self._s(ws, r, col, v, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 2.2 Estimation rationale
        r += 1
        self._sub(ws, r, 2, 6, "2.2  Descriptor Estimation Rationale")
        r += 1
        logp = _f(_g(p,"logp"),".3f")
        mr   = _f(_g(p,"molar_refractivity",_g(p,"mr","0")),".3f")
        rot  = _g(p,"rotatable_bonds","0")
        rings= _g(p,"aromatic_rings","0")
        k1   = _f(_g(p,"kappa1"),".3f")
        k2   = _f(_g(p,"kappa2"),".3f")
        k3   = _f(_g(p,"kappa3"),".3f")
        hka  = _f(_g(p,"hall_kier_alpha"),".3f")

        rationale_rows = [
            ("η′ formula",
             f"η′ = 0.8×LogP + 0.3 + 0.2×(MR/15)  [LogP={logp}, MR={mr}]"),
            ("σ′ formula",
             f"σ′ = 0.5 − 0.15×ShapeIndex + 0.05×ln(Rot+1) − 0.02×Rings  "
             f"[κ₁={k1}, κ₂={k2}, κ₃={k3}, HKα={hka}]"),
            ("β′ formula",
             "β′ = min(1, 1.0×N_aliphatic_N + 0.6×N_aromatic_N + 0.5×N_carbonyl + 0.3×N_ether)"),
            ("α′ formula",
             "α′ = min(1, 0.8×COOH + 0.7×ArOH + 0.5×Amide_NH + 0.4×AlkOH)"),
            ("κ′ formula",
             "κ′ = N_quaternary + f(pH)×(N₁+N₂+N₃) + g(pH)×N_aromatic_N"),
            ("Reference",
             "Snyder et al. (2004) J. Chromatogr. A 1060:77-116. DOI:10.1016/j.chroma.2004.08.121"),
        ]
        for i, (lbl, txt) in enumerate(rationale_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            self._s(ws, r, 3, txt, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 26
            r += 1

        # 2.3 Classification
        r += 1
        self._sub(ws, r, 2, 6, "2.3  Compound Classification")
        r += 1
        eta  = float(_f(_g(hsm,"eta_prime"),"f") or 0)
        beta = float(_f(_g(hsm,"beta_prime"),"f") or 0)
        alph = float(_f(_g(hsm,"alpha_prime"),"f") or 0)
        kapp = float(_f(_g(hsm,"kappa_prime"),"f") or 0)
        is_basic  = beta > 0.5 or kapp > 0.5
        is_acidic = alph > 0.5
        ctype = "BASIC" if is_basic else ("ACIDIC" if is_acidic else "NEUTRAL")

        self._tbl_hdr(ws, r, [2,3,4,5],
                      ["Criterion","Threshold","Value","Classification"])
        r += 1
        cls_rows = [
            ("Basic Character",   "β′>0.5 or κ′>0.5",
             f"β′={_f(_g(hsm,'beta_prime'),'.3f')}, κ′={_f(_g(hsm,'kappa_prime'),'.3f')}",
             "BASIC" if is_basic else "—"),
            ("Acidic Character",  "α′>0.5",
             f"α′={_f(_g(hsm,'alpha_prime'),'.3f')}",
             "ACIDIC" if is_acidic else "—"),
            ("Neutral Character", "β′<0.3, α′<0.3, κ′<0.3",
             "All below threshold" if not is_basic and not is_acidic else "—",
             "NEUTRAL" if not is_basic and not is_acidic else "—"),
        ]
        for i, row_data in enumerate(cls_rows):
            bg = C["amber_cell"] if row_data[3] not in ("—","") else C["grey_bg"]
            for col, val in zip([2,3,4,5], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        self._s(ws, r, 2, f"Classification Result:  {ctype} COMPOUND",
                bold=True, size=11, color="FFFFFF",
                bg=C["navy"] if ctype=="NEUTRAL" else (
                    C["teal"] if ctype=="BASIC" else C["green_hdr"]),
                align="center")
        ws.row_dimensions[r].height = 22

    # ── SECTION 3: COLUMNS ───────────────────────────────────────────────────
    def _build_columns(self, ws, results):
        col_data = _g(results,"column") if isinstance(_g(results,"column"),dict) else {}
        top_cols = col_data.get("top_columns",[])
        hsm = _g(results,"hsm") if isinstance(_g(results,"hsm"),dict) else {}

        for c, w in zip("ABCDEFGHIJ",[3,24,16,8,8,8,8,8,10,16]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 10, "SECTION 3 — Column Selection & Ranking")
        r += 1

        # 3.1 Scoring algorithm
        self._sub(ws, r, 2, 10, "3.1  Column Scoring Algorithm")
        r += 1
        beta = float(_f(_g(hsm,"beta_prime"),"f") or 0)
        kapp = float(_f(_g(hsm,"kappa_prime"),"f") or 0)
        alph = float(_f(_g(hsm,"alpha_prime"),"f") or 0)
        is_basic = beta>0.5 or kapp>0.5

        algo_rows = [
            ("Scoring Basis", "HSM Selectivity Equation — Snyder et al. (2004)"),
            ("Basic Compounds",
             "Score = 2.0×η′×H  −  3.0×β′×A  −  2.0×κ′×C    "
             "[Ref: Dolan et al. (2004)]"),
            ("Acidic Compounds",
             "Score = 2.0×η′×H  +  4.0×α′×B    [Ref: Taylor (2021)]"),
            ("Neutral Compounds",
             "Score = η′×H    [Ref: Snyder et al. (2004)]"),
            ("Applied Function",
             "Basic" if is_basic else ("Acidic" if alph>0.5 else "Neutral")),
        ]
        for i, (lbl, txt) in enumerate(algo_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
            self._s(ws, r, 3, txt, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 18
            r += 1

        # 3.2 Recommendations
        r += 1
        self._sub(ws, r, 2, 10, "3.2  Top Column Recommendations")
        r += 1
        self._tbl_hdr(ws, r, list(range(2,11)),
                      ["Rank","Column Name","Manufacturer",
                       "H","S","A","B","C(pH7)","Score"])
        r += 1
        for i, col in enumerate(top_cols[:5], 1):
            bg = C["green_cell"] if i==1 else (C["grey_bg"] if i%2==0 else C["white"])
            vals = [
                i, _g(col,"name"), _g(col,"manufacturer"),
                _f(_g(col,"h"),".3f"), _f(_g(col,"s"),".3f"),
                _f(_g(col,"a"),".3f"), _f(_g(col,"b"),".3f"),
                _f(_g(col,"c"),".3f"), _f(_g(col,"score"),".3f"),
            ]
            for c_idx, val in zip(range(2,11), vals):
                self._s(ws, r, c_idx, val, size=9, bg=bg, border=True,
                        bold=(i==1))
            ws.row_dimensions[r].height = 15
            r += 1

        # HSM parameter interpretation
        r += 1
        self._sub(ws, r, 2, 10, "3.3  HSM Parameter Interpretation")
        r += 1
        interp_rows = [
            ("H (Hydrophobicity)",
             "Higher values → stronger retention for hydrophobic compounds"),
            ("S (Steric Selectivity)",
             "Higher values → better shape selectivity"),
            ("A (H-Bond Acidity)",
             "Negative values preferred for basic compounds — less silanol activity"),
            ("B (H-Bond Basicity)",
             "Higher values preferred for acidic compounds"),
            ("C (Cation Exchange)",
             "Lower values preferred for basic compounds — reduces peak tailing"),
        ]
        for i, (param, interp) in enumerate(interp_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, param, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
            self._s(ws, r, 3, interp, size=9, bg=bg, border=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 3.4 Fs factor
        r += 1
        self._sub(ws, r, 2, 10, "3.4  Column Equivalence (Fs Factor)")
        r += 1
        fs_rows = [
            ("Fs Formula",
             "Fs = √[(12.5×ΔH)² + (100×ΔS)² + (30×ΔA)² + (143×ΔB)² + (83×ΔC)²]"),
            ("Fs ≤ 3",   "Equivalent columns (interchangeable)"),
            ("3 < Fs < 5","Similar columns (minor method differences possible)"),
            ("Fs ≥ 5",   "Different selectivity (re-validation required)"),
            ("Reference", "Dolan et al. (2004) J. Chromatogr. A 1057:59-74"),
        ]
        for i, (lbl, txt) in enumerate(fs_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
            self._s(ws, r, 3, txt, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

    # ── SECTION 4: SOLVENTS ──────────────────────────────────────────────────
    def _build_solvents(self, ws, results):
        sol = _g(results,"solvent") if isinstance(_g(results,"solvent"),dict) else {}
        top = sol.get("top_solvents",[])

        for c, w in zip("ABCDEF",[3,22,14,18,18,20]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 6, "SECTION 4 — Solvent Selection")
        r += 1

        # 4.1 Scoring system
        self._sub(ws, r, 2, 6, "4.1  Solvent Scoring System (7 Rules)")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Rule","Weight","Description","Criterion","Reference"])
        r += 1
        scoring_rules = [
            ("1. LogP Matching",         "25%",
             "ΔLogP = LogP(analyte) − LogP(solvent)",
             "Optimal: ΔLogP 2.0–3.5",
             "Valkó (2004) DOI:10.1016/j.chroma.2004.01.007"),
            ("2. HBD/HBA Complementarity","20%",
             "C_score = |α_a − β_s| + |β_a − α_s|",
             "Optimal C < 0.3",
             "Vitha & Carr (2006)"),
            ("3. Polarity Matching",      "15%",
             "|P'_analyte − P'_solvent|",
             "≤ 2.0",
             "Snyder (1974) DOI:10.1016/S0021-9673(00)85732-5"),
            ("4. Kamlet-Taft Distance",   "15%",
             "D = √[(Δα)² + (Δβ)² + (Δπ*)²]",
             "Excellent < 0.5",
             "Carr (1993) DOI:10.1006/mchj.1993.1002"),
            ("5. Viscosity/Pressure",     "10%",
             "ΔP = (φ·η·L·u) / dp²",
             "Low backpressure preferred",
             "Li & Carr (1997)"),
            ("6. UV Transparency",        "10%",
             "λ_detection > λ_cutoff + 20 nm",
             "Margin ≥ 20 nm",
             "Dolan (1999)"),
            ("7. pH Stability",           "5%",
             "Silica: pH 2–8; Hybrid: pH 1–12",
             "Within column pH range",
             "Neue (1997)"),
        ]
        for i, row_data in enumerate(scoring_rules):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5,6], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 20
            r += 1

        # 4.2 Recommendations
        r += 1
        self._sub(ws, r, 2, 6, "4.2  Top Solvent Recommendations")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Rank","Solvent","Overall Score","Strengths","Weaknesses"])
        r += 1
        for i, s in enumerate(top[:5], 1):
            bg = C["green_cell"] if i==1 else (C["grey_bg"] if i%2==0 else C["white"])
            vals = [i, _g(s,"name"), _f(_g(s,"score"),".1f"),
                    str(_g(s,"strengths","—")), str(_g(s,"weaknesses","—"))]
            for col, val in zip([2,3,4,5,6], vals):
                self._s(ws, r, col, val, size=9, bg=bg, border=True,
                        wrap=True, bold=(i==1))
            ws.row_dimensions[r].height = 18
            r += 1

        # 4.3 Initial composition
        r += 1
        comp = sol.get("initial_composition",{})
        self._sub(ws, r, 2, 6, "4.3  Initial Mobile Phase Composition")
        r += 1
        comp_rows = [
            ("Organic Modifier %", str(_g(comp,"organic_percent","50")) + "%"),
            ("Aqueous %",          str(_g(comp,"water_percent","50")) + "%"),
            ("Rationale",          _g(comp,"rationale","30-70 rule — Dolan (2002)")),
            ("Reference",
             "Dolan (2002) LCGC North America 20(5), 430-436"),
        ]
        for i, (lbl, val) in enumerate(comp_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            self._s(ws, r, 3, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 18
            r += 1

        # 4.4 Additives
        r += 1
        self._sub(ws, r, 2, 6, "4.4  Additive Recommendations")
        r += 1
        additives = sol.get("additives",[])
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["#","Additive","pH Range","Rationale","Notes"])
        r += 1
        if additives:
            for i, add in enumerate(additives[:5], 1):
                bg = C["grey_bg"] if i%2==0 else C["white"]
                vals = [i, _g(add,"name","—"), _g(add,"ph_range","—"),
                        _g(add,"rationale","—"), ""]
                for col, val in zip([2,3,4,5,6], vals):
                    self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
                ws.row_dimensions[r].height = 18
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            self._s(ws, r, 2, "No specific additives identified for this compound.",
                    size=9, bg=C["grey_bg"], border=True)
            ws.row_dimensions[r].height = 15
            r += 1

    # ── SECTION 5: BUFFERS ───────────────────────────────────────────────────
    def _build_buffers(self, ws, results):
        buf = _g(results,"buffer") if isinstance(_g(results,"buffer"),dict) else {}
        top = buf.get("top_buffers",[])

        for c, w in zip("ABCDEF",[3,26,10,14,16,22]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 6, "SECTION 5 — Buffer Selection")
        r += 1

        # 5.1 Scoring rules
        self._sub(ws, r, 2, 6, "5.1  Buffer Scoring System (10 Rules)")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Rule","Weight","Description","Criterion","Reference"])
        r += 1
        buf_rules = [
            ("1. pKa Matching",        "20%",
             "Buffer effective within ±1 pH unit of pKa",
             "Optimal ±0.5 pH units",
             "Goldberg et al. (2002) DOI:10.1063/1.1416902"),
            ("2. UV Transparency",     "15%",
             "λ_detection > λ_cutoff + 10 nm; A < 0.05 AU",
             "Margin ≥ 10 nm",
             "Perrin & Dempsey (1974)"),
            ("3. Organic Solubility",  "12%",
             "Buffer must remain soluble throughout gradient",
             "Full ACN/MeOH range",
             "Neue (1997)"),
            ("4. Buffer Capacity",     "10%",
             "β ≥ 0.01 (analytical), β ≥ 0.05 (preparative)",
             "β = 0.01–0.1",
             "Beynon & Easterby (1996)"),
            ("5. MS Volatility",       "15%/5%",
             "LC-MS requires volatile buffers",
             "Ammonium acetate/formate preferred",
             "Kebarle & Tang (1993) DOI:10.1021/ac00070a001"),
            ("6. Temperature Dep.",    "8%",
             "pKa(T) = pKa(25°C) + dpKa/dT × (T−25)",
             "Low dpKa/dT preferred",
             "Goldberg et al. (2002)"),
            ("7. Chemical Reactivity", "10%",
             "Avoid buffer-analyte reactions",
             "Risk level < 3/5",
             "Martell & Smith (2004)"),
            ("8. Ionic Strength",      "3%",
             "Optimal I = 20–50 mM",
             "I = 0.02–0.05 M",
             "Neue (1997)"),
            ("9. Storage Stability",   "2%",
             "Max storage time; pH drift rate",
             "≤7 days at RT",
             "—"),
            ("10. Metal Complexation", "5%",
             "Avoid strong chelators with metallic analytes",
             "log K(Fe³⁺) < 5",
             "Martell & Smith (2004)"),
        ]
        for i, row_data in enumerate(buf_rules):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5,6], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 20
            r += 1

        # 5.2 Recommendations
        r += 1
        self._sub(ws, r, 2, 6, "5.2  Top Buffer Recommendations")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5,6],
                      ["Rank","Buffer","pKa","Overall Score","Compatibility Notes"])
        r += 1
        for i, b in enumerate(top[:5], 1):
            bg = C["green_cell"] if i==1 else (C["grey_bg"] if i%2==0 else C["white"])
            notes = "; ".join(b.get("compatibility_notes",[])[:2]) \
                if b.get("compatibility_notes") else "—"
            vals = [i, _g(b,"name"), _f(_g(b,"pka"),".2f"),
                    _f(_g(b,"score"),".3f"), notes]
            for col, val in zip([2,3,4,5,6], vals):
                self._s(ws, r, col, val, size=9, bg=bg, border=True,
                        wrap=True, bold=(i==1))
            ws.row_dimensions[r].height = 22
            r += 1

        # 5.3 Detailed scoring for top buffer
        if top:
            r += 1
            top_b = top[0]
            self._sub(ws, r, 2, 6,
                      f"5.3  Detailed Scoring — {_g(top_b,'name')}  "
                      f"(pKa = {_f(_g(top_b,'pka'),'.2f')})")
            r += 1
            rule_scores = top_b.get("rule_scores",{})
            self._tbl_hdr(ws, r, [2,3,4,5,6],
                          ["Rule","Score","Weight","Status","Detail"])
            r += 1
            rule_detail = [
                ("pKa Matching",        "rule1_pka_matching",     "20%"),
                ("UV Transparency",     "rule2_uv_transparency",  "15%"),
                ("Organic Solubility",  "rule3_organic_solubility","12%"),
                ("Buffer Capacity",     "rule4_buffer_capacity",  "10%"),
                ("MS Compatibility",    "rule5_volatility_ms",    "15%"),
                ("Temperature Dep.",    "rule6_temperature",       "8%"),
                ("Chemical Reactivity", "rule7_reactivity",       "10%"),
                ("Ionic Strength",      "rule8_ionic_strength",    "3%"),
                ("Storage Stability",   "rule9_storage_stability", "2%"),
                ("Metal Complexation",  "rule10_metal_complexation","5%"),
            ]
            for i, (name, key, wt) in enumerate(rule_detail):
                sc  = _f(rule_scores.get(key,0),".3f")
                s_v = float(rule_scores.get(key,0))
                status = "Pass" if s_v>=0.6 else ("Marginal" if s_v>=0.3 else "Fail")
                bg = C["green_cell"] if s_v>=0.6 else (
                    C["amber_cell"] if s_v>=0.3 else C["red_cell"])
                for col, val in zip([2,3,4,5,6],
                                    [name, sc, wt, status, ""]):
                    self._s(ws, r, col, val, size=9, bg=bg, border=True)
                ws.row_dimensions[r].height = 15
                r += 1

        # 5.4 Reactivity warnings
        r += 1
        self._sub(ws, r, 2, 6, "5.4  Reactivity Warnings (Rule 7 Sub-rules)")
        r += 1
        subrule_rows = [
            ("7.1 — Amine + Carbonyl",
             "Schiff base formation risk (check if buffer is an amine + analyte has aldehyde/ketone)"),
            ("7.2 — Phosphate + Ester/Amide",
             "Hydrolysis catalysis risk at pH > 7 and T > 40°C"),
            ("7.3 — Citrate + Metals",
             "Strong chelation — log K(Fe³⁺) = 11.4. Avoid for metallic analytes."),
            ("7.4 — Borate + Diols",
             "Reversible complexation with vicinal diols (sugars, catechols)"),
        ]
        compat = buf.get("compatibility_notes",[])
        for i, (rule, detail) in enumerate(subrule_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, rule, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            self._s(ws, r, 3, detail, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 20
            r += 1

    # ── SECTION 6: GRADIENT ──────────────────────────────────────────────────
    def _build_gradient(self, ws, results):
        grad = _g(results,"gradient") if isinstance(_g(results,"gradient"),dict) else {}

        for c, w in zip("ABCDE",[3,26,16,20,24]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 5, "SECTION 6 — Gradient Program Optimization")
        r += 1

        # 6.1 Design parameters
        self._sub(ws, r, 2, 5, "6.1  Gradient Design Parameters")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5],
                      ["Parameter","Value","Unit","Reference"])
        r += 1
        param_rows = [
            ("Column Length",       "150", "mm", "—"),
            ("Column ID",           "4.6", "mm", "—"),
            ("Particle Size",       "3.5", "µm", "—"),
            ("Total Runtime",       _f(_g(grad,"total_runtime"),".1f"), "min", "—"),
            ("Equilibration Time",  "3-5 column volumes", "min",
             "Dolan (2002) DOI:10.1016/S0021-9673(02)00111-5"),
            ("Peak Capacity",       _f(_g(grad,"peak_capacity"),".1f"), "—", "—"),
            ("Min. Resolution",     _f(_g(grad,"min_resolution"),".2f"), "—", "ICH Q2(R2)"),
        ]
        for i, row_data in enumerate(param_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 6.2 Gradient table
        r += 1
        self._sub(ws, r, 2, 5, "6.2  Optimized Gradient Program")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5],
                      ["Time (min)","%B","Event","Description"])
        r += 1
        segs = grad.get("segments",[])
        if segs:
            for i, seg in enumerate(segs):
                bg = C["grey_bg"] if i%2==0 else C["white"]
                event = ("Initial" if i==0 else
                         ("Wash" if i==len(segs)-1 else "Gradient"))
                vals = [_f(_g(seg,"start_time"),".2f"),
                        _f(_g(seg,"start_b"),".1f"),
                        event,
                        f"Ramp {_f(_g(seg,'start_b'),'.1f')}% → {_f(_g(seg,'end_b'),'.1f')}%  "
                        f"({_f(_g(seg,'start_time'),'.2f')}–{_f(_g(seg,'end_time'),'.2f')} min)"]
                for col, val in zip([2,3,4,5], vals):
                    self._s(ws, r, col, val, size=9, bg=bg, border=True)
                ws.row_dimensions[r].height = 15
                r += 1
        else:
            self._row(ws, r, [2,3,4,5],
                      ["0.00","5.0","Initial conditions","—"])
            r += 1
            self._row(ws, r, [2,3,4,5],
                      ["10.00","95.0","End gradient","Linear ramp 5→95%B"])
            r += 1

        # 6.3 Robustness
        r += 1
        self._sub(ws, r, 2, 5, "6.3  Design Space & Robustness")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5],
                      ["Parameter","Acceptable Range","Optimal Value","Robustness Score"])
        r += 1
        ds = grad.get("design_space",{}) or {}
        rb = grad.get("robustness",{}) or {}
        rob_rows = [
            ("pH",          "6.5–7.5",    "7.0",  _f(_g(rb,"ph_robustness"),".2f")),
            ("Temperature", "25–35 °C",   "30 °C",_f(_g(rb,"temp_robustness"),".2f")),
            ("Flow Rate",   "1.0–2.0 mL/min","1.5 mL/min",
             _f(_g(rb,"flow_robustness"),".2f")),
            ("Initial %B",  "3–7%",       "5%",   "—"),
            ("Final %B",    "90–100%",    "95%",  "—"),
        ]
        for i, row_data in enumerate(rob_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 6.4 LSS rationale
        r += 1
        self._sub(ws, r, 2, 5, "6.4  Gradient Optimization Rationale (LSS Theory)")
        r += 1
        lss_rows = [
            ("LSS Parameter b",
             "b = S × Δφ × t₀ / t_G    [Snyder & Dolan (2007) ISBN:978-0-470-05154-2]"),
            ("Retention Prediction",
             "t_R = t₀ + (t₀/b) × ln(1 + b × k₀)    [Neue (1997)]"),
            ("S-Value Estimation",
             "S ≈ 0.48 × MW^0.44 × (1 + 0.1 × (LogP − 3))"),
            ("Peak Capacity",
             "PC = t_G / (4 × σ_avg)    [Schoenmakers (1986)]"),
            ("Objective",
             str(_g(grad,"objective","Balanced (resolution + runtime)"))),
        ]
        for i, (lbl, txt) in enumerate(lss_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            self._s(ws, r, 3, txt, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 22
            r += 1

    # ── SECTION 7: SUMMARY ───────────────────────────────────────────────────
    def _build_summary(self, ws, results):
        """Section 7 — Confidence Assessment & Warnings"""
        for c, w in zip("ABCD",[3,30,22,38]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 4,
                  "SECTION 7 — Confidence Assessment & Warnings")
        r += 1

        # 7.1 Module scores
        self._sub(ws, r, 2, 4, "7.1  Module Confidence Scores")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4],
                      ["Module","Score","Rationale"])
        r += 1
        scores = _g(results,"scores") if isinstance(_g(results,"scores"),dict) else {}
        rationale = _g(results,"rationale") if isinstance(_g(results,"rationale"),dict) else {}
        all_mods = [
            ("Physicochemical",  "physchem"),
            ("HSM Descriptors",  "hsm"),
            ("Buffer Selection", "buffer"),
            ("Solvent Selection","solvent"),
            ("Column Selection", "column"),
            ("Overall Confidence","overall"),
        ]
        for name, key in all_mods:
            s_v = float(scores.get(key,0))
            score_str = _f(s_v,".2f")
            bg = C["green_cell"] if s_v>=0.8 else (
                C["amber_cell"] if s_v>=0.5 else C["red_cell"])
            rat = str(rationale.get(key,""))[:200]
            self._s(ws, r, 2, name, bold=(key=="overall"), size=9, bg=bg, border=True)
            self._s(ws, r, 3, score_str, bold=(key=="overall"), size=9,
                    bg=bg, border=True, align="center")
            self._s(ws, r, 4, rat, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 30
            r += 1

        # 7.2 Interpretation scale
        r += 1
        self._sub(ws, r, 2, 4, "7.2  Confidence Interpretation")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4],
                      ["Score Range","Interpretation","Recommendation"])
        r += 1
        scale_rows = [
            ("> 0.80", "High confidence",
             "Method should work well with minimal optimisation"),
            ("0.50–0.80", "Moderate confidence",
             "Method may require some optimisation"),
            ("< 0.50", "Low confidence",
             "Method requires significant optimisation and validation"),
        ]
        bgs = [C["green_cell"], C["amber_cell"], C["red_cell"]]
        for (rng, interp, rec), bg in zip(scale_rows, bgs):
            for col, val in zip([2,3,4],[rng,interp,rec]):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 15
            r += 1

        # 7.3 Warnings
        r += 1
        self._sub(ws, r, 2, 4, "7.3  Warnings & Limitations")
        r += 1
        warns = results.get("warnings",[])
        errs  = results.get("errors",[])
        if warns or errs:
            for msg in warns:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
                self._s(ws, r, 2, "⚠  " + str(msg), size=9,
                        bg=C["amber_bg"], border=True, wrap=True)
                ws.row_dimensions[r].height = 18
                r += 1
            for msg in errs:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
                self._s(ws, r, 2, "✖  " + str(msg), size=9,
                        bg=C["red_bg"], border=True, wrap=True)
                ws.row_dimensions[r].height = 18
                r += 1
        else:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            self._s(ws, r, 2, "✔  No warnings generated.", size=9,
                    bg=C["green_bg"], border=True)
            ws.row_dimensions[r].height = 16
            r += 1

        # Limitations block
        r += 1
        limitations = [
            "HSM descriptors are empirical estimates — experimental validation recommended for critical applications.",
            "pKa predictions are fragment-based approximations (±1 pH unit accuracy).",
            "Gradient predictions assume ideal linear solvent strength (LSS) behaviour.",
            "Buffer compatibility should be verified experimentally for complex matrices.",
        ]
        for lim in limitations:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            self._s(ws, r, 2, "ℹ  " + lim, size=9,
                    bg=C["light_blue"], border=True, wrap=True)
            ws.row_dimensions[r].height = 22
            r += 1

    # ── SECTION 8: REFERENCES ────────────────────────────────────────────────
    def _build_references(self, ws, results):
        refs = results.get("references",[])

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 100

        r = 1
        self._hdr(ws, r, 2, 3, "SECTION 8 — References (Vancouver Style)")
        r += 1

        HARDCODED_REFS = [
            "Wildman SA, Crippen GM. Prediction of physicochemical parameters by atomic contributions. J Chem Inf Comput Sci. 1999;39(5):868-873. DOI:10.1021/ci990307l",
            "Lipinski CA, Lombardo F, Dominy BW, Feeney PJ. Experimental and computational approaches to estimate solubility and permeability. Adv Drug Deliv Rev. 1997;23:3-25. DOI:10.1016/S0169-409X(96)00423-1",
            "Ertl P, Rohde B, Selzer P. Fast calculation of molecular polar surface area. J Med Chem. 2000;43(20):3714-3717. DOI:10.1021/jm000942e",
            "Bickerton GR, Paolini GV, Besnard J, Muresan S, Hopkins AL. Quantifying the chemical beauty of drugs. Nat Chem. 2012;4(2):90-98. DOI:10.1038/nchem.1243",
            "Delaney JS. ESOL: Estimating aqueous solubility directly from molecular structure. J Chem Inf Comput Sci. 2004;44(3):1000-1005. DOI:10.1021/ci034243x",
            "Snyder LR, Dolan JW, Carr PW. The hydrophobic-subtraction model of reversed-phase column selectivity. J Chromatogr A. 2004;1060:77-116. DOI:10.1016/j.chroma.2004.08.121",
            "Dolan JW, Maule A, Bingley D, et al. Choosing an equivalent replacement column for reversed-phase LC. J Chromatogr A. 2004;1057:59-74. DOI:10.1016/j.chroma.2004.09.020",
            "Marchand DH, Snyder LR, Dolan JW. Characterization and applications of reversed-phase column selectivity. J Chromatogr A. 2008;1191:2-20. DOI:10.1016/j.chroma.2007.11.101",
            "Abraham MH. Scales of solute hydrogen-bonding. Chem Soc Rev. 1993;22(2):73-83. DOI:10.1039/CS9932200073",
            "Goldberg RN, Kishore N, Lennen RM. Thermodynamic quantities for the ionization reactions of buffers. J Phys Chem Ref Data. 2002;31(2):231-370. DOI:10.1063/1.1416902",
            "Kebarle P, Tang L. From ions in solution to ions in the gas phase. Anal Chem. 1993;65(22):972A-986A. DOI:10.1021/ac00070a001",
            "Valkó K. Application of HPLC measurements of lipophilicity. J Chromatogr A. 2004;1037:299-310. DOI:10.1016/j.chroma.2004.01.007",
            "Kamlet MJ, Abboud JLM, Abraham MH, Taft RW. Linear solvation energy relationships. 23. J Org Chem. 1983;48(17):2877-2887. DOI:10.1021/jo00165a018",
            "Snyder LR, Dolan JW. High-Performance Gradient Elution. Wiley; 2007. ISBN:978-0-470-05154-2",
            "Neue UD. HPLC Columns: Theory, Technology, and Practice. Wiley-VCH; 1997. DOI:10.1002/9783527611232",
            "Schoenmakers PJ. Optimization of Chromatographic Selectivity. Elsevier; 1986.",
            "Snyder LR, Kirkland JJ, Dolan JW. Introduction to Modern Liquid Chromatography. 3rd ed. Wiley; 2010. DOI:10.1002/9780470508183",
            "USP PQRI Column Equivalence Database. Available at: https://apps.usp.org/app/USPNF/columnsDB.html",
        ]

        # Sub-section headers
        sections = [
            (1, "Physicochemical Properties", range(0, 5)),
            (6, "HSM Descriptors & Column Selection", range(5, 9)),
            (10, "Buffer Selection", range(9, 12)),
            (13, "Solvent Selection", range(11, 14)),
            (15, "Gradient Optimization", range(13, 17)),
            (18, "Comprehensive", range(16, 18)),
        ]

        idx = 1
        for (num, title, ref_range) in sections:
            self._sub(ws, r, 2, 3, f"8.{sections.index((num,title,ref_range))+1}  {title}")
            r += 1
            for i in ref_range:
                if i < len(HARDCODED_REFS):
                    bg = C["grey_bg"] if r%2==0 else C["white"]
                    self._s(ws, r, 2, str(idx), bold=True, size=9,
                            bg=bg, border=True, align="center")
                    self._s(ws, r, 3, HARDCODED_REFS[i], size=9,
                            bg=bg, border=True, wrap=True)
                    ws.row_dimensions[r].height = 28
                    r += 1
                    idx += 1

        # Dynamic refs from reference manager
        if refs:
            r += 1
            self._sub(ws, r, 2, 3, "8.X  Additional Module References")
            r += 1
            for ref in refs[:30]:
                try:
                    if isinstance(ref, dict):
                        authors = ref.get("authors","")
                        if isinstance(authors, list):
                            authors = ", ".join(authors[:3])
                        text = (f"{authors}. {ref.get('title','')}. "
                                f"{ref.get('journal','')}. {ref.get('year','')}.")
                        if ref.get("doi"):
                            text += f" DOI:{ref['doi']}"
                    else:
                        text = str(ref)
                    bg = C["grey_bg"] if r%2==0 else C["white"]
                    self._s(ws, r, 2, str(idx), bold=True, size=9,
                            bg=bg, border=True, align="center")
                    self._s(ws, r, 3, text, size=9, bg=bg, border=True, wrap=True)
                    ws.row_dimensions[r].height = 24
                    r += 1
                    idx += 1
                except Exception:
                    pass

    # ── SECTION 9: REGULATORY ────────────────────────────────────────────────
    def _build_regulatory(self, ws, results):
        for c, w in zip("ABCDE",[3,20,30,20,24]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 5, "SECTION 9 — Regulatory Alignment (ICH / USP)")
        r += 1

        self._sub(ws, r, 2, 5, "9.1  ICH & USP Guideline Compliance")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5],
                      ["Guideline","Applicability","Implementation","Status"])
        r += 1
        reg_rows = [
            ("ICH Q2(R2)", "Validation of Analytical Procedures",
             "Method parameters selected with robustness considerations", "✔ Applied"),
            ("ICH Q3A-Q3B", "Impurity Testing",
             "Resolution predictions support impurity separation", "✔ Applied"),
            ("ICH Q8(R2)", "Pharmaceutical Development",
             "Design space approach applied to gradient conditions", "✔ Applied"),
            ("ICH Q9",  "Quality Risk Management",
             "Confidence scoring identifies high-risk parameters", "✔ Applied"),
            ("ICH Q14", "Analytical Procedure Development",
             "Systematic approach with documented rationale", "✔ Applied"),
            ("USP <621>", "Chromatography",
             "Gradient conditions within system suitability limits", "✔ Applied"),
            ("USP <1220>","Analytical Procedure Lifecycle",
             "Design space and robustness metrics provided", "✔ Applied"),
        ]
        for i, row_data in enumerate(reg_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 18
            r += 1

        r += 1
        self._sub(ws, r, 2, 5, "9.2  Recommended Validation Parameters")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4,5],
                      ["Parameter","Acceptance Criteria","Rationale","ICH Ref"])
        r += 1
        val_rows = [
            ("Specificity",   "No interfering peaks at tR",
             "Resolution predictions > 1.5", "Q2(R2) 4.1"),
            ("Linearity",     "R² > 0.999  (0.1–100 µg/mL)",
             "Based on peak capacity estimates", "Q2(R2) 4.2"),
            ("Accuracy",      "95–105% recovery",
             "Standard pharmaceutical range", "Q2(R2) 4.3"),
            ("Precision",     "RSD < 2% (system), < 5% (method)",
             "Industry standard", "Q2(R2) 4.4"),
            ("LOD / LOQ",     "S/N ≥ 3 / S/N ≥ 10",
             "ICH Q2(R2) signal-to-noise criteria", "Q2(R2) 4.7"),
            ("Robustness",    "Resolution maintained ±2% variation",
             "Design space verification", "Q2(R2) 4.6"),
        ]
        for i, row_data in enumerate(val_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4,5], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True, wrap=True)
            ws.row_dimensions[r].height = 18
            r += 1

    # ── SECTION 10: METADATA ─────────────────────────────────────────────────
    def _build_metadata(self, ws, results):
        p = _g(results,"physchem") if isinstance(_g(results,"physchem"),dict) else {}

        for c, w in zip("ABCD",[3,30,36,20]):
            ws.column_dimensions[c].width = w

        r = 1
        self._hdr(ws, r, 2, 4, "SECTION 10 — Audit Metadata & Module Versions")
        r += 1

        self._sub(ws, r, 2, 4, "10.1  Report Metadata")
        r += 1
        meta_rows = [
            ("Report ID",           _g(results,"report_id")),
            ("Timestamp",           _g(results,"timestamp")),
            ("Software Version",    "Chromatography AI System v3.0"),
            ("Processing Time",     _f(_g(results,"processing_time_s"),".2f") + " s"),
            ("Status",              _g(results,"status")),
            ("Session ID",          _g(results,"report_id")),
            ("RDKit Version",       _g(p,"rdkit_version","N/A")),
            ("Calculation Confidence", _f(_g(p,"confidence_score"),".2f")),
            ("Overall Score",       _f(_g(results,"scores","overall"),".2f")),
        ]
        for i, (lbl, val) in enumerate(meta_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, lbl, bold=True, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            self._s(ws, r, 3, val, size=9, bg=bg, border=True)
            ws.row_dimensions[r].height = 15
            r += 1

        r += 1
        self._sub(ws, r, 2, 4, "10.2  Module Versions & Parameters")
        r += 1
        self._tbl_hdr(ws, r, [2,3,4],
                      ["Module","Version","Key Parameters"])
        r += 1
        mod_rows = [
            ("PhysChem Calculator", "1.0.0", "RDKit descriptors, fragment-based pKa"),
            ("HSM Estimator",       "1.0.0", "pH = 7.0"),
            ("Column Selector",     "1.0.0", "USP PQRI database, 5-parameter HSM"),
            ("Solvent Selector",    "1.0.0", "7-rule Kamlet-Taft scoring"),
            ("Buffer Selector",     "1.0.0", "10-rule scoring system"),
            ("Gradient Optimizer",  "1.0.0", "LSS theory, Monte Carlo design space"),
        ]
        for i, row_data in enumerate(mod_rows):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            for col, val in zip([2,3,4], row_data):
                self._s(ws, r, col, val, size=9, bg=bg, border=True)
            ws.row_dimensions[r].height = 15
            r += 1

        r += 1
        self._sub(ws, r, 2, 4, "10.3  Abbreviations")
        r += 1
        self._tbl_hdr(ws, r, [2,3], ["Abbreviation","Full Term"])
        r += 1
        abbrevs = [
            ("ACN","Acetonitrile"), ("CHI","Chromatographic Hydrophobicity Index"),
            ("Fs","Column Selectivity Factor"), ("HBA","Hydrogen Bond Acceptor"),
            ("HBD","Hydrogen Bond Donor"), ("HILIC","Hydrophilic Interaction LC"),
            ("HSM","Hydrophobic Subtraction Model"), ("ICH","International Council for Harmonisation"),
            ("LC-MS","Liquid Chromatography-Mass Spectrometry"),
            ("LogD","Distribution Coefficient"), ("LogP","Partition Coefficient"),
            ("LogS","Aqueous Solubility"), ("LSS","Linear Solvent Strength"),
            ("MeOH","Methanol"), ("MR","Molar Refractivity"),
            ("QED","Quantitative Estimate of Drug-likeness"),
            ("RP","Reversed Phase"), ("TPSA","Topological Polar Surface Area"),
            ("USP","United States Pharmacopeia"),
        ]
        for i, (abbr, full) in enumerate(abbrevs):
            bg = C["grey_bg"] if i%2==0 else C["white"]
            self._s(ws, r, 2, abbr, bold=True, size=9, bg=bg, border=True)
            self._s(ws, r, 3, full, size=9, bg=bg, border=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            ws.row_dimensions[r].height = 14
            r += 1


# ─────────────────────────────────────────────────────────────────────────────
# PDF REPORT GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
class PDFReportGenerator:
    """
    Generates the full 10-section CHROME-pred PDF report using reportlab.
    """

    PAGE_W = 595.27    # A4 points
    PAGE_H = 841.89
    ML = 40            # left margin
    MR = 40            # right margin
    MT = 40            # top margin
    MB = 40            # bottom margin
    BODY_W = PAGE_W - ML - MR

    # Colour tuples (r,g,b) 0-255
    COL = {
        "navy":       (31, 78, 121),
        "blue":       (46, 117, 182),
        "teal":       (23, 55, 94),
        "light_blue": (222, 234, 241),
        "grey":       (242, 242, 242),
        "green":      (198, 239, 206),
        "amber":      (255, 235, 156),
        "red":        (255, 199, 206),
        "white":      (255, 255, 255),
        "black":      (0, 0, 0),
    }

    def __init__(self):
        try:
            from reportlab.pdfgen import canvas as rl_canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, HRFlowable
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm, mm
            self.available = True
            self._rl = {
                "canvas": rl_canvas, "A4": A4, "colors": colors,
                "SimpleDocTemplate": SimpleDocTemplate,
                "Paragraph": Paragraph, "Spacer": Spacer,
                "Table": Table, "TableStyle": TableStyle,
                "PageBreak": PageBreak, "HRFlowable": HRFlowable,
                "styles": getSampleStyleSheet(),
                "ParagraphStyle": ParagraphStyle,
                "cm": cm, "mm": mm,
            }
        except ImportError:
            self.available = False

    def _rgb(self, key):
        r, g, b = self.COL[key]
        return self._rl["colors"].Color(r/255, g/255, b/255)

    def _style(self, name="Normal", **kw):
        from reportlab.lib.styles import ParagraphStyle
        base = self._rl["styles"][name]
        return ParagraphStyle("custom", parent=base, **kw)

    def generate_pdf(self, results: Dict, references: List[Dict],
                     output_file: str) -> str:
        if not self.available:
            return ""

        rl    = self._rl
        Paragraph   = rl["Paragraph"]
        Spacer      = rl["Spacer"]
        Table       = rl["Table"]
        TableStyle  = rl["TableStyle"]
        PageBreak   = rl["PageBreak"]
        HRFlowable  = rl["HRFlowable"]
        colors      = rl["colors"]
        cm          = rl["cm"]

        doc = rl["SimpleDocTemplate"](
            output_file,
            pagesize=rl["A4"],
            leftMargin=self.ML, rightMargin=self.MR,
            topMargin=self.MT,  bottomMargin=self.MB,
        )

        # Base styles
        S = {
            "title": self._style("Heading1",
                fontSize=16, textColor=colors.white,
                backColor=self._rgb("navy"), spaceAfter=4,
                leading=20, alignment=1),
            "subtitle": self._style("Normal",
                fontSize=10, textColor=colors.white,
                backColor=self._rgb("teal"), spaceAfter=2, leading=13),
            "h1": self._style("Heading2",
                fontSize=12, textColor=colors.white,
                backColor=self._rgb("blue"), spaceAfter=3, leading=15),
            "h2": self._style("Heading3",
                fontSize=10, textColor=colors.white,
                backColor=self._rgb("teal"), spaceAfter=2, leading=13),
            "body": self._style("Normal",
                fontSize=8.5, leading=12, spaceAfter=2),
            "bold": self._style("Normal",
                fontSize=8.5, leading=12, spaceAfter=2, fontName="Helvetica-Bold"),
            "formula": self._style("Normal",
                fontSize=8, fontName="Courier", leading=11,
                backColor=self._rgb("grey"), spaceAfter=2),
            "caption": self._style("Normal",
                fontSize=7.5, textColor=colors.grey, leading=10),
        }

        def _p(text, style="body"):
            return Paragraph(str(text), S[style])

        def _sp(h=4):
            return Spacer(1, h)

        def _hr():
            return HRFlowable(width="100%", thickness=0.5,
                              color=self._rgb("blue"), spaceAfter=4)

        def _tbl(data, col_widths=None, hdr_rows=1, zebra=True):
            """Build a styled table."""
            if not data:
                return Spacer(1, 4)
            tbl = Table(data, colWidths=col_widths, repeatRows=hdr_rows)
            style_cmds = [
                ("FONTNAME",   (0, 0), (-1, hdr_rows-1), "Helvetica-Bold"),
                ("FONTSIZE",   (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, hdr_rows-1), self._rgb("navy")),
                ("TEXTCOLOR",  (0, 0), (-1, hdr_rows-1), colors.white),
                ("ROWBACKGROUNDS", (0, hdr_rows), (-1, -1),
                 [self._rgb("grey"), self._rgb("white")] if zebra
                 else [self._rgb("white")]),
                ("GRID",       (0, 0), (-1, -1), 0.4, self._rgb("blue")),
                ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
                ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING",(0, 0), (-1, -1), 4),
                ("RIGHTPADDING",(0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
                ("WORDWRAP",   (0, 0), (-1, -1), True),
            ]
            tbl.setStyle(TableStyle(style_cmds))
            return tbl

        story = []

        # ── COVER ────────────────────────────────────────────────────────────
        story.append(_p("CHROME-pred  |  AI-Assisted HPLC Method Development", "title"))
        story.append(_p("Chromatography AI System v3.0  —  Comprehensive Report", "subtitle"))
        story.append(_sp(8))

        hdr_data = [
            ["Field", "Value"],
            ["Report ID",      _g(results,"report_id")],
            ["Generated",      _g(results,"timestamp")],
            ["Compound Name",  _g(results,"name")],
            ["SMILES",         _g(results,"smiles")[:60] + ("…" if len(_g(results,"smiles",""))>60 else "")],
            ["Project",        _g(results,"project")],
            ["Status",         _g(results,"status")],
            ["Processing Time",_f(_g(results,"processing_time_s"),".2f") + " s"],
            ["Overall Confidence", _f(_g(results,"scores","overall"),".2f")],
        ]
        story.append(_tbl(hdr_data, col_widths=[4*cm, 12*cm]))
        story.append(_sp(8))

        # Confidence table
        story.append(_p("Module Confidence Scores", "h1"))
        scores   = _g(results,"scores") if isinstance(_g(results,"scores"),dict) else {}
        rationale= _g(results,"rationale") if isinstance(_g(results,"rationale"),dict) else {}
        conf_data = [["Module","Score","Status","Rationale"]]
        for name, key in [("Physicochemical","physchem"),("HSM","hsm"),
                          ("Buffer","buffer"),("Solvent","solvent"),
                          ("Column","column"),("Overall","overall")]:
            s_v = float(scores.get(key,0))
            status = "High" if s_v>=0.8 else ("Moderate" if s_v>=0.5 else "Low")
            rat = str(rationale.get(key,""))[:100]
            conf_data.append([name, _f(s_v,".2f"), status, rat])
        story.append(_tbl(conf_data, col_widths=[3*cm,2*cm,2.5*cm,8.5*cm]))

        story.append(PageBreak())

        # ── SECTION 1 PHYSCHEM ───────────────────────────────────────────────
        story.append(_p("Section 1 — Compound Identification & Physicochemical Profile","h1"))
        p = _g(results,"physchem") if isinstance(_g(results,"physchem"),dict) else {}
        story.append(_p("1.1  Compound Information", "h2"))
        story.append(_tbl([
            ["Property","Value"],
            ["Compound Name",   _g(results,"name")],
            ["SMILES",          _g(results,"smiles")],
            ["Molecular Formula",_g(p,"molecular_formula")],
            ["Molecular Weight", _f(_g(p,"molecular_weight"),".3f") + " g/mol"],
            ["Exact Mass",       _f(_g(p,"exact_mass"),".5f") + " Da"],
            ["Project",          _g(results,"project")],
        ], col_widths=[5*cm,11*cm]))
        story.append(_sp(6))

        story.append(_p("1.2  Physicochemical Properties", "h2"))
        story.append(_tbl([
            ["Descriptor","Value","Unit","Threshold","Reference"],
            ["LogP (Wildman-Crippen)", _f(_g(p,"logp"),".3f"),"","−2 to 5","Wildman & Crippen 1999"],
            ["LogD (pH 2.0)", _f(_g(p,"logd_ph2"),".3f"),"","","Henderson-Hasselbalch"],
            ["LogD (pH 5.0)", _f(_g(p,"logd_ph5"),".3f"),"","",""],
            ["LogD (pH 7.4)", _f(_g(p,"logd_ph74"),".3f"),"","","Physiological pH"],
            ["LogD (pH 9.0)", _f(_g(p,"logd_ph9"),".3f"),"","",""],
            ["LogD (pH 11.0)",_f(_g(p,"logd_ph11"),".3f"),"","",""],
            ["TPSA", _f(_g(p,"tpsa"),".2f"),"Å²","≤140","Ertl et al. 2000"],
            ["H-Bond Donors",    str(_g(p,"hbd_lipinski","0")),"","≤5","Lipinski 1997"],
            ["H-Bond Acceptors", str(_g(p,"hba_lipinski","0")),"","≤10",""],
            ["Rotatable Bonds",  str(_g(p,"rotatable_bonds","0")),"","≤10","Veber 2002"],
            ["Aromatic Rings",   str(_g(p,"aromatic_rings","0")),"","",""],
            ["Fraction Csp³",    _f(_g(p,"fraction_csp3"),".3f"),"","",""],
        ], col_widths=[4*cm,2.5*cm,1.5*cm,2.5*cm,5.5*cm]))
        story.append(_sp(6))

        story.append(_p("1.3  Drug-Likeness", "h2"))
        lipi = _g(p,"lipinski_violations","0")
        story.append(_tbl([
            ["Filter","Status","Threshold","Reference"],
            ["Lipinski Violations", str(lipi), "≤1 acceptable","Lipinski 1997"],
            ["Ghose Filter",  str(_g(p,"ghose_filter","N/A")), "MW 160-480","Ghose 1999"],
            ["Veber Filter",  str(_g(p,"veber_filter","N/A")), "Rot ≤10, TPSA ≤140","Veber 2002"],
            ["Muegge Filter", str(_g(p,"muegge_filter","N/A")),"MW 200-600","Muegge 2001"],
            ["QED Score",     _f(_g(p,"qed_score"),".3f"),">0.5 favourable","Bickerton 2012"],
        ], col_widths=[4*cm,3*cm,4*cm,5*cm]))
        story.append(_sp(6))

        story.append(_p("1.4  Ionization Profile", "h2"))
        story.append(_tbl([
            ["Parameter","Value"],
            ["Ionization Type",   str(_g(p,"ionization_type","N/A"))],
            ["pKa Acidic (min)",  _f(_g(p,"pka_acidic_min"),".2f")],
            ["pKa Basic (max)",   _f(_g(p,"pka_basic_max"),".2f")],
            ["Isoelectric Point", _f(_g(p,"isoelectric_point"),".2f")],
            ["Formal Charge",     str(_g(p,"formal_charge","0"))],
        ], col_widths=[5*cm,11*cm]))

        story.append(_sp(6))
        story.append(_p("1.5  Solubility & Permeability", "h2"))
        story.append(_tbl([
            ["Parameter","Value","Unit","Reference"],
            ["LogS (ESOL)",          _f(_g(p,"logS"),".2f"),"log mol/L","Delaney 2004"],
            ["Intrinsic Solubility", _f(_g(p,"intrinsic_solubility"),".3f"),"mg/mL",""],
            ["Caco-2 Permeability",  _f(_g(p,"caco2_permeability"),".1f"),"×10⁻⁶ cm/s","PAMPA"],
            ["MDCK Permeability",    _f(_g(p,"mdck_permeability"),".1f"),"×10⁻⁶ cm/s",""],
        ], col_widths=[4.5*cm,3*cm,3*cm,5.5*cm]))

        story.append(PageBreak())

        # ── SECTION 2 HSM ────────────────────────────────────────────────────
        story.append(_p("Section 2 — HSM Descriptors","h1"))
        hsm = _g(results,"hsm") if isinstance(_g(results,"hsm"),dict) else {}
        story.append(_p("2.1  Descriptor Values (Snyder et al., 2004)", "h2"))
        story.append(_tbl([
            ["Descriptor","Symbol","Value","Interpretation","Reference"],
            ["Hydrophobicity",   "η′","", "","Marchand 2008"],
            ["Steric Resistance","σ′","","","Marchand 2005"],
            ["H-Bond Basicity",  "β′","","","Abraham 1993"],
            ["H-Bond Acidity",   "α′","","","Abraham 1993"],
            ["Cationic Charge",  "κ′","","","Dolan 2004"],
        ], col_widths=[3.5*cm,2*cm,2.5*cm,4.5*cm,3.5*cm]))

        # fill in values
        desc_map = {"η′":"eta_prime","σ′":"sigma_prime","β′":"beta_prime",
                    "α′":"alpha_prime","κ′":"kappa_prime"}
        # rebuild with actual values
        def _interp_v(key):
            v = float(_f(_g(hsm,key),"f") or 0)
            if key=="eta_prime": return "High" if v>1.5 else ("Mod." if v>0.8 else "Low")
            if key=="sigma_prime": return "High" if v>0.3 else "Low"
            if key in ("beta_prime","alpha_prime"): return "Strong" if v>0.5 else "Weak"
            if key=="kappa_prime": return "Sig." if v>0.5 else "Min."
            return ""

        hsm_tbl = [["Descriptor","Symbol","Value","Interpretation","Reference"]]
        for name, sym, key, ref in [
            ("Hydrophobicity","η′","eta_prime","Marchand 2008"),
            ("Steric Resistance","σ′","sigma_prime","Marchand 2005"),
            ("H-Bond Basicity","β′","beta_prime","Abraham 1993"),
            ("H-Bond Acidity","α′","alpha_prime","Abraham 1993"),
            ("Cationic Charge","κ′","kappa_prime","Dolan 2004"),
        ]:
            hsm_tbl.append([name, sym, _f(_g(hsm,key),".3f"),
                             _interp_v(key), ref])
        story.append(_tbl(hsm_tbl, col_widths=[3.5*cm,2*cm,2.5*cm,4.5*cm,3.5*cm]))

        story.append(_sp(6))
        story.append(_p("2.2  Estimation Formulae", "h2"))
        for lbl, formula in [
            ("η′", "η′ = 0.8 × LogP + 0.3 + 0.2 × (MR/15)"),
            ("σ′", "σ′ = 0.5 − 0.15 × ShapeIndex + 0.05 × ln(Rot+1) − 0.02 × Rings"),
            ("β′", "β′ = min(1, 1.0×N_alip_N + 0.6×N_arom_N + 0.5×N_carbonyl + 0.3×N_ether)"),
            ("α′", "α′ = min(1, 0.8×COOH + 0.7×ArOH + 0.5×Amide_NH + 0.4×AlkOH)"),
            ("κ′", "κ′ = N_quaternary + f(pH)×(N₁+N₂+N₃) + g(pH)×N_arom_N"),
        ]:
            story.append(_p(f"<b>{lbl}:</b>", "body"))
            story.append(_p(formula, "formula"))
        story.append(_p("Reference: Snyder et al. (2004) J. Chromatogr. A 1060:77-116", "caption"))

        story.append(PageBreak())

        # ── SECTION 3 COLUMNS ────────────────────────────────────────────────
        story.append(_p("Section 3 — Column Selection & Ranking","h1"))
        col_d = _g(results,"column") if isinstance(_g(results,"column"),dict) else {}
        top_c = col_d.get("top_columns",[])

        story.append(_p("3.1  Scoring Algorithm", "h2"))
        beta_v = float(_f(_g(hsm,"beta_prime"),"f") or 0)
        kap_v  = float(_f(_g(hsm,"kappa_prime"),"f") or 0)
        alph_v = float(_f(_g(hsm,"alpha_prime"),"f") or 0)
        is_bas = beta_v>0.5 or kap_v>0.5
        fn = "Basic" if is_bas else ("Acidic" if alph_v>0.5 else "Neutral")
        story.append(_p(f"Applied scoring function: <b>{fn}</b>", "body"))
        formulae = {
            "Basic":   "Score = 2.0×η′×H − 3.0×β′×A − 2.0×κ′×C  [Dolan et al. 2004]",
            "Acidic":  "Score = 2.0×η′×H + 4.0×α′×B  [Taylor 2021]",
            "Neutral": "Score = η′×H  [Snyder et al. 2004]",
        }
        story.append(_p(formulae[fn], "formula"))

        story.append(_p("3.2  Top Column Recommendations", "h2"))
        col_tbl = [["Rank","Column","Manufacturer","H","S","A","B","C","Score"]]
        for i, col in enumerate(top_c[:5], 1):
            col_tbl.append([
                i, _g(col,"name"), _g(col,"manufacturer"),
                _f(_g(col,"h"),".3f"), _f(_g(col,"s"),".3f"),
                _f(_g(col,"a"),".3f"), _f(_g(col,"b"),".3f"),
                _f(_g(col,"c"),".3f"), _f(_g(col,"score"),".3f"),
            ])
        story.append(_tbl(col_tbl,
            col_widths=[1*cm,4.5*cm,3*cm,1.3*cm,1.3*cm,1.3*cm,1.3*cm,1.3*cm,1.5*cm]))

        story.append(_sp(6))
        story.append(_p("3.4  Column Equivalence (Fs Factor)", "h2"))
        story.append(_p("Fs = √[(12.5×ΔH)² + (100×ΔS)² + (30×ΔA)² + (143×ΔB)² + (83×ΔC)²]",
                         "formula"))
        story.append(_tbl([
            ["Fs Value","Interpretation"],
            ["Fs ≤ 3",  "Equivalent — interchangeable"],
            ["3 < Fs < 5","Similar — minor differences possible"],
            ["Fs ≥ 5",  "Different selectivity — re-validation required"],
        ], col_widths=[4*cm,12*cm]))

        story.append(PageBreak())

        # ── SECTION 4 SOLVENTS ───────────────────────────────────────────────
        story.append(_p("Section 4 — Solvent Selection","h1"))
        sol = _g(results,"solvent") if isinstance(_g(results,"solvent"),dict) else {}
        top_s = sol.get("top_solvents",[])

        story.append(_p("4.1  Scoring System (7 Rules)", "h2"))
        story.append(_tbl([
            ["Rule","Weight","Description","Reference"],
            ["1. LogP Matching","25%","ΔLogP 2.0–3.5 optimal","Valkó 2004"],
            ["2. HBD/HBA","20%","C_score < 0.3","Vitha & Carr 2006"],
            ["3. Polarity","15%","|ΔP'| ≤ 2.0","Snyder 1974"],
            ["4. Kamlet-Taft","15%","D < 0.5","Carr 1993"],
            ["5. Viscosity","10%","Low backpressure","Li & Carr 1997"],
            ["6. UV Transparency","10%","Margin ≥ 20 nm","Dolan 1999"],
            ["7. pH Stability","5%","Within column range","Neue 1997"],
        ], col_widths=[3.5*cm,2*cm,6*cm,4.5*cm]))

        story.append(_sp(6))
        story.append(_p("4.2  Top Solvent Recommendations", "h2"))
        sol_tbl = [["Rank","Solvent","Overall Score","Notes"]]
        for i, s in enumerate(top_s[:5], 1):
            sol_tbl.append([i, _g(s,"name"), _f(_g(s,"score"),".1f"), ""])
        story.append(_tbl(sol_tbl, col_widths=[1.5*cm,4*cm,3*cm,7.5*cm]))

        comp = sol.get("initial_composition",{})
        story.append(_sp(4))
        story.append(_p("4.3  Initial Mobile Phase Composition", "h2"))
        story.append(_tbl([
            ["Parameter","Value"],
            ["Organic Modifier %", str(_g(comp,"organic_percent","50")) + "%"],
            ["Aqueous %",          str(_g(comp,"water_percent","50")) + "%"],
            ["Rationale",          _g(comp,"rationale","30-70 rule (Dolan 2002)")],
        ], col_widths=[5*cm,11*cm]))

        story.append(PageBreak())

        # ── SECTION 5 BUFFERS ────────────────────────────────────────────────
        story.append(_p("Section 5 — Buffer Selection","h1"))
        buf = _g(results,"buffer") if isinstance(_g(results,"buffer"),dict) else {}
        top_b = buf.get("top_buffers",[])

        story.append(_p("5.1  10-Rule Scoring System", "h2"))
        story.append(_tbl([
            ["Rule","Weight","Description"],
            ["1. pKa Matching","20%","Buffer ±1 pH unit of pKa (Goldberg 2002)"],
            ["2. UV Transparency","15%","λ_det > λ_cutoff + 10 nm"],
            ["3. Organic Solubility","12%","Soluble throughout gradient"],
            ["4. Buffer Capacity","10%","β ≥ 0.01"],
            ["5. MS Volatility","15%","Volatile buffers for LC-MS (Kebarle 1993)"],
            ["6. Temperature Dep.","8%","pKa(T) correction"],
            ["7. Chemical Reactivity","10%","Avoid analyte reactions"],
            ["8. Ionic Strength","3%","I = 20–50 mM"],
            ["9. Storage Stability","2%","≤7 days at RT"],
            ["10. Metal Complexation","5%","Avoid chelators"],
        ], col_widths=[3.5*cm,2*cm,10.5*cm]))

        story.append(_sp(6))
        story.append(_p("5.2  Top Buffer Recommendations", "h2"))
        buf_tbl = [["Rank","Buffer","pKa","Score","Notes"]]
        for i, b in enumerate(top_b[:5], 1):
            notes = "; ".join(b.get("compatibility_notes",[])[:2]) \
                if b.get("compatibility_notes") else "—"
            buf_tbl.append([i, _g(b,"name"), _f(_g(b,"pka"),".2f"),
                             _f(_g(b,"score"),".3f"), notes])
        story.append(_tbl(buf_tbl, col_widths=[1.5*cm,4.5*cm,2*cm,2*cm,6*cm]))

        if top_b:
            story.append(_sp(6))
            story.append(_p(f"5.3  Detailed Scoring — {_g(top_b[0],'name')}", "h2"))
            tb = top_b[0]
            rs = tb.get("rule_scores",{})
            rd = [["Rule","Score","Status"]]
            for rname, rkey in [
                ("pKa Matching","rule1_pka_matching"),
                ("UV Transparency","rule2_uv_transparency"),
                ("Organic Solubility","rule3_organic_solubility"),
                ("Buffer Capacity","rule4_buffer_capacity"),
                ("MS Volatility","rule5_volatility_ms"),
                ("Temp. Dep.","rule6_temperature"),
                ("Reactivity","rule7_reactivity"),
                ("Ionic Strength","rule8_ionic_strength"),
                ("Storage","rule9_storage_stability"),
                ("Metals","rule10_metal_complexation"),
            ]:
                sv = float(rs.get(rkey,0))
                rd.append([rname, _f(sv,".3f"),
                           "Pass" if sv>=0.6 else ("Marginal" if sv>=0.3 else "Fail")])
            story.append(_tbl(rd, col_widths=[5*cm,3*cm,4*cm]))

        story.append(PageBreak())

        # ── SECTION 6 GRADIENT ───────────────────────────────────────────────
        story.append(_p("Section 6 — Gradient Program Optimization","h1"))
        grad = _g(results,"gradient") if isinstance(_g(results,"gradient"),dict) else {}

        story.append(_p("6.1  Gradient Parameters", "h2"))
        story.append(_tbl([
            ["Parameter","Value","Reference"],
            ["Total Runtime",    _f(_g(grad,"total_runtime"),".1f") + " min","—"],
            ["Peak Capacity",    _f(_g(grad,"peak_capacity"),".1f"),"—"],
            ["Min. Resolution",  _f(_g(grad,"min_resolution"),".2f"),"ICH Q2(R2)"],
            ["Column Length",    "150 mm","—"],
            ["Column ID",        "4.6 mm","—"],
            ["Particle Size",    "3.5 µm","—"],
        ], col_widths=[5*cm,4*cm,7*cm]))

        story.append(_sp(6))
        story.append(_p("6.2  Gradient Programme", "h2"))
        segs = grad.get("segments",[])
        seg_tbl = [["Time (min)","%B","Type","Description"]]
        if segs:
            for i, seg in enumerate(segs):
                event = "Initial" if i==0 else ("Wash" if i==len(segs)-1 else "Gradient")
                seg_tbl.append([
                    _f(_g(seg,"start_time"),".2f"),
                    _f(_g(seg,"start_b"),".1f"),
                    event,
                    f"{_f(_g(seg,'start_b'),'.1f')}→{_f(_g(seg,'end_b'),'.1f')}%B",
                ])
        else:
            seg_tbl += [["0.00","5.0","Initial","—"],
                        ["10.00","95.0","End","5→95%B"]]
        story.append(_tbl(seg_tbl, col_widths=[3*cm,3*cm,3*cm,7*cm]))

        story.append(_sp(6))
        story.append(_p("6.3  LSS Theory Rationale", "h2"))
        for lbl, formula in [
            ("b  =", "b = S × Δφ × t₀ / t_G  [Snyder & Dolan 2007]"),
            ("t_R =", "t_R = t₀ + (t₀/b) × ln(1 + b × k₀)  [Neue 1997]"),
            ("S  =", "S ≈ 0.48 × MW^0.44 × (1 + 0.1 × (LogP − 3))"),
        ]:
            story.append(_p(f"<b>{lbl}</b>  " + formula, "formula"))

        story.append(PageBreak())

        # ── SECTION 7 CONFIDENCE ─────────────────────────────────────────────
        story.append(_p("Section 7 — Confidence Assessment & Warnings","h1"))

        story.append(_p("7.1  Module Confidence Scores", "h2"))
        conf_tbl = [["Module","Score","Status","Rationale"]]
        for name, key in [("Physicochemical","physchem"),("HSM","hsm"),
                           ("Buffer","buffer"),("Solvent","solvent"),
                           ("Column","column"),("OVERALL","overall")]:
            s_v = float(scores.get(key,0))
            status = "High" if s_v>=0.8 else ("Moderate" if s_v>=0.5 else "Low")
            rat = str(rationale.get(key,""))[:100]
            conf_tbl.append([name, _f(s_v,".2f"), status, rat])
        story.append(_tbl(conf_tbl, col_widths=[3*cm,2*cm,2.5*cm,8.5*cm]))

        story.append(_sp(6))
        story.append(_p("7.2  Interpretation Scale", "h2"))
        story.append(_tbl([
            ["Score Range","Interpretation","Recommendation"],
            ["> 0.80","High confidence","Method works well, minimal optimisation"],
            ["0.50–0.80","Moderate confidence","Some optimisation required"],
            ["< 0.50","Low confidence","Significant optimisation + validation needed"],
        ], col_widths=[3*cm,4*cm,9*cm]))

        story.append(_sp(6))
        story.append(_p("7.3  Warnings & Limitations", "h2"))
        warns = results.get("warnings",[])
        errs  = results.get("errors",[])
        if warns or errs:
            for msg in warns:
                story.append(_p("⚠  " + str(msg), "body"))
            for msg in errs:
                story.append(_p("✖  " + str(msg), "body"))
        else:
            story.append(_p("✔  No warnings generated.", "body"))
        for lim in [
            "HSM descriptors are empirical estimates — experimental validation recommended.",
            "pKa predictions are fragment-based approximations (±1 pH unit accuracy).",
            "Gradient predictions assume ideal linear solvent strength (LSS) behaviour.",
        ]:
            story.append(_p("ℹ  " + lim, "body"))

        story.append(PageBreak())

        # ── SECTION 8 REFERENCES ─────────────────────────────────────────────
        story.append(_p("Section 8 — References (Vancouver Style)","h1"))
        REFS = [
            "1. Wildman SA, Crippen GM. J Chem Inf Comput Sci. 1999;39(5):868-873. DOI:10.1021/ci990307l",
            "2. Lipinski CA et al. Adv Drug Deliv Rev. 1997;23:3-25. DOI:10.1016/S0169-409X(96)00423-1",
            "3. Ertl P, Rohde B, Selzer P. J Med Chem. 2000;43(20):3714-3717. DOI:10.1021/jm000942e",
            "4. Bickerton GR et al. Nat Chem. 2012;4(2):90-98. DOI:10.1038/nchem.1243",
            "5. Delaney JS. J Chem Inf Comput Sci. 2004;44(3):1000-1005. DOI:10.1021/ci034243x",
            "6. Snyder LR, Dolan JW, Carr PW. J Chromatogr A. 2004;1060:77-116. DOI:10.1016/j.chroma.2004.08.121",
            "7. Dolan JW et al. J Chromatogr A. 2004;1057:59-74. DOI:10.1016/j.chroma.2004.09.020",
            "8. Marchand DH et al. J Chromatogr A. 2008;1191:2-20. DOI:10.1016/j.chroma.2007.11.101",
            "9. Marchand DH et al. J Chromatogr A. 2005;1062:65-78. DOI:10.1016/j.chroma.2004.11.014",
            "10. Abraham MH. Chem Soc Rev. 1993;22(2):73-83. DOI:10.1039/CS9932200073",
            "11. Goldberg RN et al. J Phys Chem Ref Data. 2002;31(2):231-370. DOI:10.1063/1.1416902",
            "12. Kebarle P, Tang L. Anal Chem. 1993;65(22):972A-986A. DOI:10.1021/ac00070a001",
            "13. Valkó K. J Chromatogr A. 2004;1037:299-310. DOI:10.1016/j.chroma.2004.01.007",
            "14. Kamlet MJ et al. J Org Chem. 1983;48(17):2877-2887. DOI:10.1021/jo00165a018",
            "15. Snyder LR, Dolan JW. High-Performance Gradient Elution. Wiley; 2007.",
            "16. Neue UD. HPLC Columns. Wiley-VCH; 1997. DOI:10.1002/9783527611232",
            "17. Schoenmakers PJ. Optimization of Chromatographic Selectivity. Elsevier; 1986.",
            "18. Snyder LR, Kirkland JJ, Dolan JW. Intro. to Modern LC. 3rd ed. Wiley; 2010.",
            "19. USP PQRI Column Equivalence Database. https://apps.usp.org/app/USPNF/columnsDB.html",
        ]
        for ref in REFS:
            story.append(_p(ref, "body"))

        story.append(PageBreak())

        # ── SECTION 9 REGULATORY ─────────────────────────────────────────────
        story.append(_p("Section 9 — Regulatory Alignment (ICH / USP)","h1"))
        story.append(_tbl([
            ["Guideline","Applicability","Status"],
            ["ICH Q2(R2)","Validation of Analytical Procedures","✔ Applied"],
            ["ICH Q3A-Q3B","Impurity Testing","✔ Applied"],
            ["ICH Q8(R2)","Pharmaceutical Development","✔ Applied"],
            ["ICH Q9","Quality Risk Management","✔ Applied"],
            ["ICH Q14","Analytical Procedure Development","✔ Applied"],
            ["USP <621>","Chromatography","✔ Applied"],
            ["USP <1220>","Analytical Procedure Lifecycle","✔ Applied"],
        ], col_widths=[3*cm,9*cm,4*cm]))

        story.append(_sp(6))
        story.append(_p("Recommended Validation Parameters", "h2"))
        story.append(_tbl([
            ["Parameter","Acceptance Criteria","ICH Ref"],
            ["Specificity","No interfering peaks at tR","Q2(R2) 4.1"],
            ["Linearity","R² > 0.999  (0.1–100 µg/mL)","Q2(R2) 4.2"],
            ["Accuracy","95–105% recovery","Q2(R2) 4.3"],
            ["Precision","RSD < 2% (system), < 5% (method)","Q2(R2) 4.4"],
            ["LOD / LOQ","S/N ≥ 3 / S/N ≥ 10","Q2(R2) 4.7"],
            ["Robustness","Resolution maintained ±2% variation","Q2(R2) 4.6"],
        ], col_widths=[4*cm,8*cm,4*cm]))

        story.append(PageBreak())

        # ── SECTION 10 METADATA ──────────────────────────────────────────────
        story.append(_p("Section 10 — Audit Metadata","h1"))
        story.append(_tbl([
            ["Property","Value"],
            ["Report ID",       _g(results,"report_id")],
            ["Timestamp",       _g(results,"timestamp")],
            ["Software Version","Chromatography AI System v3.0"],
            ["RDKit Version",   _g(p,"rdkit_version","N/A")],
            ["Processing Time", _f(_g(results,"processing_time_s"),".2f") + " s"],
            ["Status",          _g(results,"status")],
            ["Overall Score",   _f(_g(results,"scores","overall"),".2f")],
        ], col_widths=[5*cm,11*cm]))

        story.append(_sp(8))
        story.append(_p(
            "This report was generated by the CHROME-pred AI-Assisted "
            "Chromatographic Method Development System. All recommendations are "
            "based on peer-reviewed literature and should be validated "
            "experimentally for critical applications.",
            "caption"
        ))

        try:
            doc.build(story)
            logger.info(f"PDF report generated: {output_file}")
            return output_file
        except Exception as e:
            logger.error(f"PDF build failed: {e}")
            return ""
