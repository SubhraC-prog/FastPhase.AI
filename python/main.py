#!/usr/bin/env python3
"""
AI-Assisted Chromatographic Method Development System v2.0
Integrated with Method_Development_Input.xlsm

This script reads from the Excel input file, processes compounds through all
analysis modules with full reference tracking, and writes results back.

Excel Structure:
- Sheet1: "Input" (Main Interface with SMILES entry, batch queue, live preview)
- Sheet2: "History" (Recent compounds with timestamps)
- Sheet3: "Templates" (Saved method templates)
- Sheet4: "Settings" (User preferences)
- Sheet5: "Help" (Documentation)
- Sheet6: "About" (Version info)
"""

import os
import sys
import time
import logging
import warnings
import traceback
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple, Union
from enum import Enum
from dataclasses import asdict
import json
import hashlib
import uuid

# Data handling
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

# Suppress warnings
warnings.filterwarnings('ignore')

# Configure logging
# File handler records everything; stream handler only shows errors to avoid
# spurious stderr output which the Excel macro treats as a failure.
file_handler = logging.FileHandler(Path(__file__).parent / 'chromatography_ai.log')
stream_handler = logging.StreamHandler()
stream_handler.setLevel(logging.ERROR)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        file_handler,
        stream_handler
    ]
)
logger = logging.getLogger(__name__)

# Import all modules with reference tracking
try:
    from HSMsolute_check import HSMEstimator
    HAS_HSM = True
except ImportError as e:
    HAS_HSM = False
    logger.warning(f"HSM Estimator not available: {e}")

try:
    from column_selector import HSMColumnSelector, HSMReference
    HAS_COLUMN = True
except ImportError as e:
    HAS_COLUMN = False
    logger.warning(f"Column Selector not available: {e}")

try:
    from solvent_selector import SolventSelectionSystem, ReferenceCollector
    HAS_SOLVENT = True
except ImportError as e:
    HAS_SOLVENT = False
    logger.warning(f"Solvent Selector not available: {e}")

try:
    # module file is named `buffer_selector.py` (lowercase); import using correct name
    from buffer_selector import BufferSelector, BufferSelectionEngine
    HAS_BUFFER = True
except ImportError as e:
    HAS_BUFFER = False
    logger.warning(f"Buffer Selector not available: {e}")

try:
    from gradient_optimizer import GradientOptimizer, OptimizationObjective, GradientType
    HAS_GRADIENT = True
except ImportError as e:
    HAS_GRADIENT = False
    logger.warning(f"Gradient Optimizer not available: {e}")

try:
    from physchem_calculator import PhysicochemicalCalculator, IonizationType
    HAS_PHYSCHEM = True
except ImportError as e:
    HAS_PHYSCHEM = False
    logger.warning(f"Physicochemical Calculator not available: {e}")

# Import report template and Excel formatting
try:
    from report_template import ReportTemplate, PDFReportGenerator
    HAS_REPORT_TEMPLATE = True
except ImportError as e:
    HAS_REPORT_TEMPLATE = False
    logger.warning(f"Report Template not available: {e}")

try:
    from excel_formatter import ExcelStyleManager
    HAS_EXCEL_FORMATTER = True
except ImportError as e:
    HAS_EXCEL_FORMATTER = False
    logger.warning(f"Excel Formatter not available: {e}")

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class SimplePDFReportGenerator:
    """Minimal fallback PDF generator using reportlab."""
    def generate_pdf(self, results: Dict, references: List[Dict], output_file: str) -> str:
        try:
            c = canvas.Canvas(output_file, pagesize=letter)
            width, height = letter
            margin = inch * 0.75
            text = c.beginText(margin, height - margin)
            text.setFont('Helvetica-Bold', 14)
            text.textLine('CHROME-pred Report')
            text.setFont('Helvetica', 10)
            text.textLine(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            text.textLine('')
            text.setFont('Helvetica-Bold', 12)
            text.textLine('Summary')
            text.setFont('Helvetica', 10)

            summary_items = [
                ('Report ID', results.get('report_id', 'N/A')),
                ('Compound Name', results.get('name', 'N/A')),
                ('Project', results.get('project', 'N/A')),
                ('SMILES', results.get('smiles', 'N/A')),
                ('Status', results.get('status', 'N/A')),
                ('Overall Confidence', str(results.get('scores', {}).get('overall', 'N/A'))),
                ('Processing Time (s)', str(results.get('processing_time_s', 'N/A')))
            ]
            for label, value in summary_items:
                text.textLine(f"{label}: {value}")

            text.textLine('')
            text.setFont('Helvetica-Bold', 12)
            text.textLine('Top Recommendations')
            text.setFont('Helvetica', 10)

            if 'column' in results and results['column'].get('top_columns'):
                top_column = results['column']['top_columns'][0]
                text.textLine(f"Column: {top_column.get('name', 'N/A')} ({top_column.get('score', 0)})")
            if 'solvent' in results and results['solvent'].get('top_solvents'):
                top_solvent = results['solvent']['top_solvents'][0]
                text.textLine(f"Solvent: {top_solvent.get('name', 'N/A')} ({top_solvent.get('score', 0)})")
            if 'buffer' in results and results['buffer'].get('top_buffers'):
                top_buffer = results['buffer']['top_buffers'][0]
                text.textLine(f"Buffer: {top_buffer.get('name', 'N/A')} ({top_buffer.get('score', 0)})")
            if 'gradient' in results:
                text.textLine(f"Gradient Runtime: {results['gradient'].get('total_runtime', 'N/A')} min")

            text.textLine('')
            text.setFont('Helvetica-Bold', 12)
            text.textLine('Warnings / Errors')
            text.setFont('Helvetica', 10)
            for entry in results.get('warnings', []):
                text.textLine(f"Warning: {entry}")
            for entry in results.get('errors', []):
                text.textLine(f"Error: {entry}")

            text.textLine('')
            text.setFont('Helvetica-Bold', 12)
            text.textLine('References')
            text.setFont('Helvetica', 10)
            for idx, ref in enumerate(references[:10], 1):
                if isinstance(ref, dict):
                    citation = ref.get('title', '') or ref.get('citation', '')
                    text.textLine(f"{idx}. {citation}")
                else:
                    text.textLine(f"{idx}. {str(ref)}")
                if text.getY() < margin + inch:
                    c.drawText(text)
                    c.showPage()
                    text = c.beginText(margin, height - margin)
                    text.setFont('Helvetica', 10)

            c.drawText(text)
            c.save()
            return output_file
        except Exception as e:
            logger.error(f"Simple PDF generator failed: {e}")
            return ''


class ReferenceManager:
    """Manages all references from all modules with deduplication."""
    
    def __init__(self):
        self.references = {}
        self.module_refs = {}
        
    def add_reference(self, module: str, ref_key: str, ref_data: Dict):
        """Add a reference with deduplication."""
        ref_id = hashlib.md5(str(ref_data).encode()).hexdigest()
        
        if ref_id not in self.references:
            self.references[ref_id] = {
                'id': ref_id,
                'module': module,
                'key': ref_key,
                **ref_data
            }
        
        if module not in self.module_refs:
            self.module_refs[module] = []
        if ref_id not in self.module_refs[module]:
            self.module_refs[module].append(ref_id)
    
    def get_all_references(self) -> List[Dict]:
        """Get all unique references."""
        return list(self.references.values())
    
    def get_references_by_module(self, module: str) -> List[Dict]:
        """Get references for a specific module."""
        ref_ids = self.module_refs.get(module, [])
        return [self.references[rid] for rid in ref_ids if rid in self.references]
    
    def format_citations(self, style: str = 'apa') -> str:
        """Format all references in specified citation style."""
        citations = []
        for ref in self.get_all_references():
            if style == 'apa':
                if 'authors' in ref:
                    citations.append(f"{ref['authors']} ({ref.get('year', 'n.d.')}). {ref.get('title', '')}. {ref.get('journal', '')}, {ref.get('volume', '')}, {ref.get('pages', '')}.")
                elif 'database' in ref:
                    citations.append(f"{ref['database']}. Available at: {ref.get('url', '')}")
            elif style == 'acs':
                if 'authors' in ref:
                    citations.append(f"{ref['authors']} {ref.get('journal', '')} {ref.get('year', '')}, {ref.get('volume', '')}, {ref.get('pages', '')}.")
        return '\n'.join(citations)


class ChromatographyAIController:
    """
    Main controller for the Chromatography AI system.
    Follows the exact Excel structure from Method_Development_Input.xlsm.
    """
    
    # Excel cell positions based on the provided structure
    EXCEL_CELLS = {
        'smiles_input': 'B2',
        'compound_name': 'B3',
        'project': 'B4',
        'notes': 'B5',
        'generate_report_btn': 'E2',
        'batch_process_btn': 'F2',
        'load_example_btn': 'G2',
        'validate_smiles_btn': 'E3',
        'clear_all_btn': 'F3',
        'save_template_btn': 'G3',
        'method_wizard_btn': 'E4',
        'export_input_btn': 'F4',
        'load_template_btn': 'G4',
        'preferences_btn': 'E5',
        'view_logs_btn': 'F5',
        'help_btn': 'G5',
        'advanced_options_btn': 'E6',
        'reset_all_btn': 'F6',
        'report_bug_btn': 'G6',
        'status': 'H2',
        'validation': 'H3',
        'last_run': 'H4',
        'progress': 'H5',
        'eta': 'H6',
        'logp': 'A7',
        'mol_weight': 'B7',
        'tpsa': 'A8',
        'hbd_hba': 'B8',
        'rotatable': 'A9',
        'rings': 'B9',
        'formula': 'A10',
        'violations': 'B10',
        'open_last_report': 'E9',
        'compare_methods': 'F9',
        'view_dashboard': 'G9',
        'batch_queue_start': 14,  # Row 14 is start of batch queue
    }
    
    def __init__(self):
        """Initialize the controller and all available modules."""
        logger.info("=" * 80)
        logger.info("CHROMATOGRAPHY AI SYSTEM v2.0")
        logger.info("=" * 80)
        
        # Initialize modules
        self.modules = {}
        self.reference_manager = ReferenceManager()
        self._initialize_modules()
        
        # Excel file paths
        self.excel_file = None
        self.output_dir = Path(__file__).parent / "output"
        self.output_dir.mkdir(exist_ok=True)
        self.reports_dir = self.output_dir / "reports"
        self.reports_dir.mkdir(exist_ok=True)
        self.logs_dir = self.output_dir / "logs"
        self.logs_dir.mkdir(exist_ok=True)
        
        # Current state
        self.current_smiles = None
        self.current_results = None
        self.batch_queue = []
        self.history = []
        
        # Load settings if available
        self.settings = self._load_default_settings()
        
        logger.info("System initialization complete")
        self._print_module_status()
    
    def _print_module_status(self):
        """Print status of available modules."""
        logger.info("\nModule Status:")
        logger.info(f"  HSM Estimator: {'Available' if HAS_HSM else 'Unavailable'}")
        logger.info(f"  Column Selector: {'Available' if HAS_COLUMN else 'Unavailable'}")
        logger.info(f"  Solvent Selector: {'Available' if HAS_SOLVENT else 'Unavailable'}")
        logger.info(f"  Buffer Selector: {'Available' if HAS_BUFFER else 'Unavailable'}")
        logger.info(f"  Gradient Optimizer: {'Available' if HAS_GRADIENT else 'Unavailable'}")
        logger.info(f"  PhysChem Calculator: {'Available' if HAS_PHYSCHEM else 'Unavailable'}")
    
    def _load_default_settings(self) -> Dict:
        """Load default settings."""
        return {
            'python_path': 'python',
            'timeout': 30,
            'auto_save': True,
            'debug_mode': False,
            'output_folder': str(self.output_dir),
            'auto_open_pdf': True,
            'include_references': True,
            'include_regulatory': True,
            'buffer_model': '2.0.0',
            'solvent_model': '2.0.0',
            'column_model': '2.0.0',
            'confidence_threshold': 0.70,
            'theme': 'Professional',
            'font_size': 11,
            'show_live_preview': True
        }
    
    def _initialize_modules(self):
        """Initialize all available modules and collect their references."""
        
        # HSM Estimator
        if HAS_HSM:
            try:
                self.modules['hsm'] = HSMEstimator(pH=7.0)
                logger.info("  HSM Estimator initialized")
                
                # Add HSM references
                self.reference_manager.add_reference('HSM', 'SNYDER_2004', {
                    'authors': 'Snyder, L. R., Dolan, J. W., & Carr, P. W.',
                    'title': 'The hydrophobic-subtraction model of reversed-phase column selectivity',
                    'journal': 'Journal of Chromatography A',
                    'volume': '1060',
                    'pages': '77-116',
                    'year': 2004,
                    'doi': '10.1016/j.chroma.2004.08.121'
                })
                self.reference_manager.add_reference('HSM', 'ABRAHAM_1993', {
                    'authors': 'Abraham, M. H.',
                    'title': 'Scales of solute hydrogen-bonding: their construction and application',
                    'journal': 'Chemical Society Reviews',
                    'volume': '22',
                    'pages': '73-83',
                    'year': 1993
                })
            except Exception as e:
                logger.error(f"  HSM Estimator failed: {e}")
        
        # Column Selector
        if HAS_COLUMN:
            try:
                self.modules['column'] = HSMColumnSelector(pH=7.0)
                logger.info("  Column Selector initialized")
                
                # Add column selector references
                self.reference_manager.add_reference('Column', 'DOLAN_2004', {
                    'authors': 'Dolan, J. W., Maule, A., Bingley, D., et al.',
                    'title': 'Choosing an equivalent replacement column for reversed-phase LC',
                    'journal': 'Journal of Chromatography A',
                    'volume': '1057',
                    'pages': '59-74',
                    'year': 2004
                })
                self.reference_manager.add_reference('Column', 'MARCHAND_2005', {
                    'authors': 'Marchand, D. H., et al.',
                    'title': 'Column selectivity in reversed-phase liquid chromatography. VIII',
                    'journal': 'Journal of Chromatography A',
                    'volume': '1062',
                    'pages': '65-78',
                    'year': 2005
                })
                self.reference_manager.add_reference('Column', 'USP_PQRI', {
                    'database': 'USP PQRI Column Equivalence Database',
                    'url': 'https://apps.usp.org/app/USPNF/columnsDB.html'
                })
            except Exception as e:
                logger.error(f"  Column Selector failed: {e}")
        
        # Solvent Selector
        if HAS_SOLVENT:
            try:
                self.modules['solvent'] = SolventSelectionSystem()
                logger.info("  Solvent Selector initialized")
                
                # Add solvent selector references
                self.reference_manager.add_reference('Solvent', 'VALKO_2004', {
                    'authors': 'Valkó, K.',
                    'title': 'Application of HPLC measurements of lipophilicity to model biological distribution',
                    'journal': 'Journal of Chromatography A',
                    'volume': '1037',
                    'pages': '299-310',
                    'year': 2004
                })
                self.reference_manager.add_reference('Solvent', 'KAMLET_1983', {
                    'authors': 'Kamlet, M. J., Abboud, J. L. M., Abraham, M. H., & Taft, R. W.',
                    'title': 'Linear solvation energy relationships. 23. A comprehensive collection',
                    'journal': 'Journal of Organic Chemistry',
                    'volume': '48',
                    'pages': '2877-2887',
                    'year': 1983
                })
                self.reference_manager.add_reference('Solvent', 'SNYDER_1974', {
                    'authors': 'Snyder, L. R.',
                    'title': 'Classification of the solvent properties of common liquids',
                    'journal': 'Journal of Chromatography A',
                    'volume': '92',
                    'pages': '223-230',
                    'year': 1974
                })
            except Exception as e:
                logger.error(f"  Solvent Selector failed: {e}")
        
        # Buffer Selector
        if HAS_BUFFER:
            try:
                self.modules['buffer'] = BufferSelector()
                logger.info("  Buffer Selector initialized")
                
                # Add buffer selector references
                self.reference_manager.add_reference('Buffer', 'GOLDBERG_2002', {
                    'authors': 'Goldberg, R. N., Kishore, N., & Lennen, R. M.',
                    'title': 'Thermodynamic quantities for the ionization reactions of buffers',
                    'journal': 'Journal of Physical and Chemical Reference Data',
                    'volume': '31',
                    'pages': '231-370',
                    'year': 2002
                })
                self.reference_manager.add_reference('Buffer', 'PERRIN_1974', {
                    'authors': 'Perrin, D. D., & Dempsey, B.',
                    'title': 'Buffers for pH and Metal Ion Control',
                    'journal': 'Chapman and Hall',
                    'year': 1974
                })
                self.reference_manager.add_reference('Buffer', 'KEBARLE_1993', {
                    'authors': 'Kebarle, P., & Tang, L.',
                    'title': 'From ions in solution to ions in the gas phase',
                    'journal': 'Analytical Chemistry',
                    'volume': '65',
                    'pages': '972A-986A',
                    'year': 1993
                })
            except Exception as e:
                logger.error(f"  Buffer Selector failed: {e}")
        
        # Gradient Optimizer
        if HAS_GRADIENT:
            try:
                self.modules['gradient'] = GradientOptimizer(
                    column_length=150,
                    column_id=4.6,
                    particle_size=3.5
                )
                logger.info("  Gradient Optimizer initialized")
                
                # Add gradient optimizer references
                self.reference_manager.add_reference('Gradient', 'SNYDER_2007', {
                    'authors': 'Snyder, L. R., & Dolan, J. W.',
                    'title': 'High-Performance Gradient Elution',
                    'journal': 'Wiley',
                    'year': 2007
                })
                self.reference_manager.add_reference('Gradient', 'NEUE_1997', {
                    'authors': 'Neue, U. D.',
                    'title': 'HPLC Columns: Theory, Technology, and Practice',
                    'journal': 'Wiley-VCH',
                    'year': 1997
                })
                self.reference_manager.add_reference('Gradient', 'SCHOENMAKERS_1986', {
                    'authors': 'Schoenmakers, P. J.',
                    'title': 'Optimization of Chromatographic Selectivity',
                    'journal': 'Elsevier',
                    'year': 1986
                })
            except Exception as e:
                logger.error(f"  Gradient Optimizer failed: {e}")
        
        # Physicochemical Calculator
        if HAS_PHYSCHEM:
            logger.info("  Physicochemical Calculator ready")
    
    def load_excel_file(self, filepath: Union[str, Path]) -> bool:
        """
        Load the Method_Development_Input.xlsm file.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            bool: True if successful
        """
        self.excel_file = Path(filepath)
        
        if not self.excel_file.exists():
            logger.error(f"Excel file not found: {self.excel_file}")
            return False
        
        logger.info(f"Loaded Excel file: {self.excel_file}")
        
        try:
            # Ensure required sheets exist (for new workbooks)
            wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)
            self._ensure_master_workbook_structure(wb)
            wb.save(self.excel_file)

            # Load settings from Settings sheet if available
            self._load_settings_from_excel()
            
            # Load history from History sheet
            self._load_history_from_excel()
            
            return True
            
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return False
    
    def _load_settings_from_excel(self):
        """Load user preferences from Settings sheet."""
        try:
            df_settings = pd.read_excel(self.excel_file, sheet_name='Settings', header=None)
            
            # Parse settings (assuming format from your structure)
            for _, row in df_settings.iterrows():
                if len(row) >= 3 and pd.notna(row[1]):
                    category = str(row[0]) if pd.notna(row[0]) else ''
                    param = str(row[1]) if pd.notna(row[1]) else ''
                    value = row[2] if pd.notna(row[2]) else ''
                    
                    # Map to settings
                    if 'Python Interpreter' in param:
                        self.settings['python_path'] = str(value)
                    elif 'Timeout' in param:
                        self.settings['timeout'] = int(value) if str(value).isdigit() else 30
                    elif 'Auto-save' in param:
                        self.settings['auto_save'] = str(value).lower() == 'yes'
                    elif 'Debug Mode' in param:
                        self.settings['debug_mode'] = str(value).lower() == 'yes'
                    elif 'Default Output Folder' in param:
                        self.settings['output_folder'] = str(value)
                    elif 'Include References' in param:
                        self.settings['include_references'] = str(value).lower() == 'yes'
            
            logger.info("Settings loaded from Excel")
            
        except Exception as e:
            logger.warning(f"Could not load settings: {e}")
    
    def _load_history_from_excel(self):
        """Load recent compounds from History sheet."""
        try:
            # Prefer the new "Search History" sheet if available
            sheet_name = 'Search History'
            try:
                df_history = pd.read_excel(self.excel_file, sheet_name=sheet_name)
            except Exception:
                sheet_name = 'History'
                df_history = pd.read_excel(self.excel_file, sheet_name=sheet_name)

            if not df_history.empty:
                self.history = df_history.to_dict('records')
                logger.info(f"Loaded {len(self.history)} history entries from {sheet_name} sheet")

        except Exception as e:
            logger.warning(f"Could not load history: {e}")

    def _ensure_master_workbook_structure(self, wb):
        """Ensure the master Excel workbook has all expected sheets and headers."""
        required_sheets = [
            'Input', 'History', 'Search History', 'Settings',
            'Templates', 'Help', 'About'
        ]
        for sheet in required_sheets:
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)

        # Ensure Search History has proper headers
        self._ensure_search_history_sheet(wb)

    def _ensure_search_history_sheet(self, wb):
        """Ensure the Search History sheet exists and has the correct header row."""
        sheet_name = 'Search History'
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # Create headers if missing
        if ws.max_row < 1 or ws.cell(row=1, column=1).value != 'Timestamp':
            headers = [
                'Timestamp', 'Report ID', 'SMILES', 'Compound Name', 'Project',
                'Status', 'Overall Confidence', 'PhysChem Score', 'HSM Score',
                'Buffer Score', 'Solvent Score', 'Column Score', 'Processing Time (s)',
                'Excel Report', 'PDF Report', 'Raw Data'
            ]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

        # Hide raw JSON column
        ws.column_dimensions['P'].hidden = True

        # Conditional formatting for status
        status_col = 'F'
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}1048576",
            CellIsRule(operator='equal', formula=['"Success"'], stopIfTrue=True,
                       fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}1048576",
            CellIsRule(operator='equal', formula=['"Partial"'], stopIfTrue=True,
                       fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'))
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}1048576",
            CellIsRule(operator='equal', formula=['"Failed"'], stopIfTrue=True,
                       fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
        )

        # Confidence score color scale (Cols H-L)
        for col in range(8, 13):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}1048576",
                ColorScaleRule(start_type='num', start_value=0, start_color='FFC7CE',
                               mid_type='num', mid_value=0.6, mid_color='FFEB9C',
                               end_type='num', end_value=0.8, end_color='C6EFCE')
            )

        return ws

    def _get_existing_sheet_name(self, candidates: list, default: str) -> str:
        """Return the first existing sheet name from candidates, else default."""
        if not self.excel_file:
            return default
        try:
            wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)
            for name in candidates:
                if name in wb.sheetnames:
                    return name
        except Exception:
            pass
        return default

    def read_input_smiles(self) -> Tuple[str, str, str, str]:
        """
        Read SMILES and related fields from the Input/Dashboard sheet.
        
        Returns:
            Tuple of (smiles, compound_name, project, notes)
        """
        if not self.excel_file:
            return "", "", "", ""
        
        try:
            sheet_name = self._get_existing_sheet_name(['DASHBOARD', 'Input'], 'Input')
            df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=None)
            
            # Map to cell positions
            cell_map = {
                'C5': (4, 2),  # SMILES Input  — DASHBOARD C5 (0-indexed row 4, col 2)
                'C6': (5, 2),  # Compound Name — DASHBOARD C6
                'C7': (6, 2),  # Project       — DASHBOARD C7
                'C8': (7, 2),  # Notes         — DASHBOARD C8
            }
            
            smiles = ""
            name = ""
            project = ""
            notes = ""
            
            for cell_name, (row, col) in cell_map.items():
                if row < len(df) and col < len(df.columns):
                    value = df.iat[row, col]
                    if pd.notna(value):
                        if cell_name == 'C5':
                            smiles = str(value).strip()
                        elif cell_name == 'C6':
                            name = str(value).strip()
                        elif cell_name == 'C7':
                            project = str(value).strip()
                        elif cell_name == 'C8':
                            notes = str(value).strip()
            
            self.current_smiles = smiles
            logger.info(f"Read SMILES: {smiles[:50]}...")
            
            return smiles, name, project, notes
            
        except Exception as e:
            logger.error(f"Error reading input: {e}")
            return "", "", "", ""
    
    def read_batch_queue(self) -> List[Dict]:
        """
        Read batch queue from Input sheet starting at row 14.
        
        Returns:
            List of batch queue entries
        """
        if not self.excel_file:
            return []
        
        try:
            df = pd.read_excel(self.excel_file, sheet_name='Input', header=None)
            
            batch_queue = []
            start_row = 13  # 0-indexed row 13 = Excel row 14
            
            for i in range(start_row, len(df)):
                row = df.iloc[i]
                if len(row) >= 3 and pd.notna(row[1]):  # Check SMILES column
                    entry = {
                        'row': i + 1,  # Excel row number
                        'number': row[0] if pd.notna(row[0]) else '',
                        'smiles': str(row[1]).strip() if pd.notna(row[1]) else '',
                        'name': str(row[2]).strip() if pd.notna(row[2]) else '',
                        'status': str(row[3]).strip() if len(row) > 3 and pd.notna(row[3]) else 'Pending',
                        'progress': str(row[4]).strip() if len(row) > 4 and pd.notna(row[4]) else ''
                    }
                    if entry['smiles']:
                        batch_queue.append(entry)
                else:
                    break  # Stop at first empty row
            
            self.batch_queue = batch_queue
            logger.info(f"Read {len(batch_queue)} batch queue entries")
            
            return batch_queue
            
        except Exception as e:
            logger.error(f"Error reading batch queue: {e}")
            return []
    
    def update_live_preview(self, smiles: str):
        """
        Update live preview area (cells A7-B10) with calculated properties.
        
        Args:
            smiles: SMILES string to calculate properties for
        """
        if not self.excel_file or not smiles:
            return
        
        try:
            # Calculate properties
            properties = self._calculate_properties(smiles)
            
            # Load workbook
            wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)
            ws_name = self._get_existing_sheet_name(['DASHBOARD', 'Input'], 'DASHBOARD')
            ws = wb[ws_name]

            # Update live preview cells
            ws['A7'] = f"● LogP: {properties.get('logp', 'N/A')}"
            ws['B7'] = f"● Mol Weight: {properties.get('mw', 'N/A')}"
            ws['A8'] = f"● TPSA: {properties.get('tpsa', 'N/A')}"
            ws['B8'] = f"● HBD/HBA: {properties.get('hbd', 0)}/{properties.get('hba', 0)}"
            ws['A9'] = f"● Rotatable: {properties.get('rotatable', 0)}"
            ws['B9'] = f"● Rings: {properties.get('rings', 0)}"
            ws['A10'] = f"● Formula: {properties.get('formula', 'N/A')}"
            ws['B10'] = f"● Violations: {properties.get('violations', 0)}"
            
            # Update validation status
            ws['H3'] = "✅ Valid" if properties.get('valid', False) else "❌ Invalid"
            
            # Save workbook
            wb.save(self.excel_file)
            logger.info("Live preview updated")
            
        except Exception as e:
            logger.error(f"Error updating live preview: {e}")
    
    def _calculate_properties(self, smiles: str) -> Dict:
        """Calculate properties for live preview."""
        properties = {
            'logp': 'N/A',
            'mw': 'N/A',
            'tpsa': 'N/A',
            'hbd': 0,
            'hba': 0,
            'rotatable': 0,
            'rings': 0,
            'formula': 'N/A',
            'violations': 0,
            'valid': False
        }
        
        if HAS_PHYSCHEM:
            try:
                calc = PhysicochemicalCalculator(smiles)
                props = calc.calculate_all()
                
                properties.update({
                    'logp': f"{props.logp:.2f}",
                    'mw': f"{props.molecular_weight:.1f}",
                    'tpsa': f"{props.tpsa:.1f}",
                    'hbd': props.hbd_lipinski,
                    'hba': props.hba_lipinski,
                    'rotatable': props.rotatable_bonds,
                    'rings': props.aromatic_rings,
                    'formula': props.molecular_formula,
                    'violations': props.lipinski_violations,
                    'valid': True
                })
            except:
                pass
        
        return properties
    
    def validate_smiles(self, smiles: str) -> Tuple[bool, str]:
        """
        Validate SMILES string.
        
        Args:
            smiles: SMILES string to validate
            
        Returns:
            Tuple of (is_valid, message)
        """
        if not smiles:
            return False, "Empty SMILES string"
        
        try:
            if HAS_PHYSCHEM:
                calc = PhysicochemicalCalculator(smiles)
                calc.calculate_all()
                return True, "Valid SMILES"
            else:
                # Basic validation
                from rdkit import Chem
                mol = Chem.MolFromSmiles(smiles)
                if mol:
                    return True, "Valid SMILES"
                else:
                    return False, "Invalid SMILES"
        except Exception as e:
            return False, str(e)
    
    def process_single_compound(self, smiles: str, name: str = "", 
                                project: str = "", notes: str = "") -> Dict:
        """
        Process a single compound through all modules.
        
        Args:
            smiles: SMILES string
            name: Compound name
            project: Project name
            notes: Additional notes
            
        Returns:
            Dictionary with all results
        """
        logger.info(f"Processing compound: {name or smiles[:30]}...")
        
        start_time = time.perf_counter()
        report_id = f"RPT-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:6]}"

        results = {
            'report_id': report_id,
            'smiles': smiles,
            'name': name,
            'project': project,
            'notes': notes,
            'timestamp': datetime.now().isoformat(),
            'status': 'Success',
            'warnings': [],
            'errors': []
        }
        
        try:
            # 1. Physicochemical Properties
            if HAS_PHYSCHEM:
                try:
                    calc = PhysicochemicalCalculator(smiles)
                    props = calc.calculate_all()

                    # Convert dataclass to dict for easy reporting; fall back to __dict__ for non-dataclass objects
                    try:
                        physchem_dict = asdict(props)
                    except Exception:
                        physchem_dict = vars(props) if hasattr(props, '__dict__') else {}

                    # Ensure Enums and complex objects are stringified
                    for k, v in list(physchem_dict.items()):
                        if isinstance(v, (Enum,)):
                            physchem_dict[k] = str(v)

                    results['physchem'] = physchem_dict

                    # Add chromatographic descriptors (if available)
                    results['chrom_descriptors'] = {
                        'hydrophobic_index': getattr(props, 'hydrophobic_index', None),
                        'hydrophilic_index': getattr(props, 'hydrophilic_index', None),
                        'chrom_hydrophobicity': getattr(props, 'chromatographic_hydrophobicity', None),
                        'silanol_potential': getattr(props, 'silanol_interaction_potential', None),
                        'pi_pi_potential': getattr(props, 'pi_pi_interaction_potential', None),
                        'steric_bulk': getattr(props, 'steric_bulk_parameter', None),
                        'hbond_potential': getattr(props, 'hydrogen_bonding_potential', None),
                    }

                    # Add reference citations for primary PhysChem rules used in this report
                    self.reference_manager.add_reference('PhysChem', 'Crippen_1999', {
                        'authors': 'Crippen, G. M. & Wildman, S. A.',
                        'title': 'Prediction of Hydrophobicity from Molecular Structure',
                        'journal': 'J. Chem. Inf. Comput. Sci.',
                        'year': 1999,
                        'doi': '10.1021/ci980426i'
                    })
                    self.reference_manager.add_reference('PhysChem', 'Lipinski_1997', {
                        'authors': 'Lipinski, C. A.',
                        'title': 'Experimental and computational approaches to estimate solubility',
                        'journal': 'Adv. Drug Deliv. Rev.',
                        'year': 1997,
                        'doi': '10.1016/S0169-409X(98)00118-5'
                    })
                    self.reference_manager.add_reference('PhysChem', 'Ertl_2000', {
                        'authors': 'Ertl, P., Rohde, B., & Selzer, P.',
                        'title': 'Fast calculation of molecular polar surface area',
                        'journal': 'J. Med. Chem.',
                        'year': 2000,
                        'doi': '10.1021/jm000942e'
                    })

                except Exception as e:
                    results['warnings'].append(f"PhysChem: {str(e)}")
            
            # 2. HSM Descriptor Estimation
            if HAS_HSM and 'hsm' in self.modules:
                try:
                    hsm = self.modules['hsm']
                    descriptors = hsm.calculate_from_smiles(smiles)
                    
                    results['hsm'] = {
                        'eta_prime': descriptors.get('η_prime', 0),
                        'sigma_prime': descriptors.get('σ_prime', 0),
                        'beta_prime': descriptors.get('β_prime', 0),
                        'alpha_prime': descriptors.get('α_prime', 0),
                        'kappa_prime': descriptors.get('κ_prime', 0),
                    }
                    
                    # Add references used
                    if hasattr(hsm, 'get_references_used'):
                        refs = hsm.get_references_used()
                        for ref in refs:
                            self.reference_manager.add_reference('HSM', 'custom', {'citation': ref})
                    
                except Exception as e:
                    results['warnings'].append(f"HSM: {str(e)}")
            
            # 3. Column Selection
            if HAS_COLUMN and 'column' in self.modules:
                try:
                    column = self.modules['column']
                    
                    # Use HSM descriptors if available
                    if 'hsm' in results:
                        column.last_descriptors = {
                            'η_prime': results['hsm']['eta_prime'],
                            'σ_prime': results['hsm']['sigma_prime'],
                            'β_prime': results['hsm']['beta_prime'],
                            'α_prime': results['hsm']['alpha_prime'],
                            'κ_prime': results['hsm']['kappa_prime'],
                        }
                    
                    recommendations = column.select_columns_for_smiles(
                        smiles, n_recommendations=5, show_references=False
                    )
                    
                    results['column'] = {
                        'top_columns': []
                    }
                    
                    for _, row in recommendations.head(5).iterrows():
                        results['column']['top_columns'].append({
                            'name': row.get('Name', ''),
                            'manufacturer': row.get('Manufacturer', ''),
                            'h': row.get('H', 0),
                            's': row.get('S', 0),
                            'a': row.get('A', 0),
                            'b': row.get('B', 0),
                            'c': row.get('C_pH7', 0),
                            'score': row.get('Score', 0)
                        })
                    
                    # Get advice
                    advice = column.get_selection_advice()
                    results['column']['advice'] = advice
                    
                    # Add references
                    if hasattr(column.scoring_engine, 'references_used'):
                        for ref in column.scoring_engine.references_used:
                            self.reference_manager.add_reference('Column', 'custom', {'citation': ref})
                    
                except Exception as e:
                    results['warnings'].append(f"Column: {str(e)}")
            
            # 4. Solvent Selection
            if HAS_SOLVENT and 'solvent' in self.modules:
                try:
                    solvent = self.modules['solvent']
                    
                    # Use default parameters
                    solvent_results = solvent.process_single_smiles(
                        smiles,
                        name=name,
                        detection_wavelength=254,
                        ph=7.0,
                        temperature=30,
                        column_type='silica'
                    )
                    
                    if solvent_results:
                        results['solvent'] = {
                            'top_solvents': []
                        }
                        
                        for rec in solvent_results.get('recommendations', [])[:5]:
                            results['solvent']['top_solvents'].append({
                                'name': rec.get('solvent', ''),
                                'score': rec.get('total_score', 0)
                            })
                        
                        composition = solvent_results.get('initial_composition', {})
                        results['solvent']['initial_composition'] = {
                            'organic_percent': composition.get('organic_percent', 50),
                            'water_percent': composition.get('water_percent', 50),
                            'rationale': composition.get('rationale', '')
                        }
                        
                        additives = solvent_results.get('additive_recommendations', {})
                        results['solvent']['additives'] = additives.get('recommendations', [])
                    
                except Exception as e:
                    results['warnings'].append(f"Solvent: {str(e)}")
            
            # 5. Buffer Selection
            if HAS_BUFFER and 'buffer' in self.modules:
                try:
                    buffer = self.modules['buffer']
                    
                    # Prepare method parameters
                    method_params = {
                        'target_ph': 7.0,
                        'detection_wavelength_nm': 254,
                        'is_lcms': False,
                        'organic_modifier': 'ACN',
                        'max_organic_percent': 80,
                        'temperature_c': 25,
                        'buffer_concentration_mM': 50,
                        'is_hilic': False,
                        'is_preparative': False,
                        'contains_metals': False,
                        'contains_esters': False,
                        'contains_aldehydes': False,
                        'contains_diols': False,
                        'storage_required_days': 7,
                        'gradient_elution': True,
                        'max_allowable_absorbance': 0.05
                    }
                    
                    buffer.engine.set_method_parameters(**method_params)
                    buffer_results = buffer.engine.select_optimal_buffer(smiles)
                    
                    results['buffer'] = {
                        'top_buffers': [],
                        'compatibility_notes': buffer_results.get('compatibility_notes', []),
                        'method_parameters': buffer_results.get('method_parameters', {}),
                        'all_buffers': buffer_results.get('all_buffers', [])
                    }
                    
                    for buf in buffer_results.get('top_buffers', [])[:5]:
                        results['buffer']['top_buffers'].append({
                            'name': buf['base_name'],
                            'pka': buf['pka'],
                            'score': buf['final_score'],
                            'compatibility_notes': buf.get('compatibility_notes', []),
                            'rule_scores': buf.get('scores', {})
                        })
                    
                except Exception as e:
                    results['warnings'].append(f"Buffer: {str(e)}")
            
            # 6. Gradient Optimization
            if HAS_GRADIENT and 'gradient' in self.modules:
                try:
                    gradient = self.modules['gradient']
                    
                    # Prepare compound data
                    compound = {
                        'name': name or "Compound",
                        'logp': results.get('physchem', {}).get('logp', 3.0),
                        'molecular_weight': results.get('physchem', {}).get('molecular_weight', 300),
                        'tpsa': results.get('physchem', {}).get('tpsa', 50),
                    }
                    
                    constraints = {
                        'max_time': 60,
                        'min_resolution': 1.5,
                        'max_pressure': 400,
                        'min_b': 5,
                        'max_b': 95
                    }
                    
                    opt_result = gradient.optimize_gradient(
                        compounds=[compound],
                        objective=OptimizationObjective.BALANCED,
                        gradient_type=GradientType.LINEAR,
                        constraints=constraints
                    )
                    
                    results['gradient'] = {
                        'total_runtime': opt_result.gradient_program.total_runtime,
                        'peak_capacity': opt_result.summary().get('peak_capacity', 0),
                        'min_resolution': opt_result.summary().get('min_resolution', 0),
                        'objective': opt_result.objective_value,
                        'confidence': opt_result.confidence_score,
                        'warnings': opt_result.warnings,
                        'robustness': opt_result.robustness_metrics,
                        'design_space': opt_result.design_space,
                        'segments': []
                    }
                    
                    for seg in opt_result.gradient_program.segments:
                        results['gradient']['segments'].append({
                            'start_time': seg.start_time,
                            'end_time': seg.end_time,
                            'start_b': seg.start_b,
                            'end_b': seg.end_b,
                            'type': seg.type.value if hasattr(seg, 'type') else 'linear',
                            'curve_factor': getattr(seg, 'curve_factor', 1.0)
                        })
                    
                except Exception as e:
                    results['warnings'].append(f"Gradient: {str(e)}")
            
            # Provide fallback selections when modules unavailable
            if not HAS_SOLVENT and 'solvent' not in results:
                results['solvent'] = {
                    'top_solvents': [{'name': 'ACN', 'score': 50}],
                    'initial_composition': {'organic_percent': 50, 'water_percent': 50, 'rationale': 'Default fallback 50/50 ACN/H2O'},
                    'additives': []
                }
            if not HAS_BUFFER and 'buffer' not in results:
                results['buffer'] = {
                    'top_buffers': [{'name': 'Phosphate 50 mM', 'pka': 7.2, 'score': 50}]
                }
            if not HAS_GRADIENT and 'gradient' not in results:
                results['gradient'] = {
                    'total_runtime': 10,
                    'initial_composition': '5%',
                    'final_composition': '95%',
                    'segments': [{'start_time': 0.0, 'end_time': 10.0, 'start_b': 5, 'end_b': 95}]
                }
            # Add to history
            self._add_to_history(results)
            
        except Exception as e:
            results['status'] = 'Failed'
            results['errors'].append(str(e))
            logger.error(f"Error processing compound: {e}")
        
        # Calculate processing time and confidence scores
        end_time = time.perf_counter()
        results['processing_time_s'] = round(end_time - start_time, 2)
        self._compute_confidence_scores(results)
        
        self.current_results = results
        return results
    
    def _compute_confidence_scores(self, results: Dict):
        """Compute confidence/score metrics for each module and overall."""
        scores = {}
        # Physicochemical confidence (use QED + Lipinski violations when available)
        phys = results.get('physchem', {})
        qed = phys.get('qed_score')
        lipinski_violations = phys.get('lipinski_violations')
        phys_score = 0.0
        count = 0
        if isinstance(qed, (int, float)):
            phys_score += float(qed)
            count += 1
        if isinstance(lipinski_violations, (int, float)):
            phys_score += max(0.0, 1.0 - min(5, float(lipinski_violations)) / 5.0)
            count += 1
        scores['physchem'] = round((phys_score / count) if count else 0.0, 2)

        # HSM confidence: average of descriptor values (expected 0-1 range)
        hsm = results.get('hsm', {})
        hsm_vals = [hsm.get(k, 0) for k in ('eta_prime', 'sigma_prime', 'beta_prime', 'alpha_prime', 'kappa_prime')]
        hsm_vals = [min(1.0, max(0.0, float(v))) for v in hsm_vals if isinstance(v, (int, float))]
        scores['hsm'] = round((sum(hsm_vals) / len(hsm_vals)) if hsm_vals else 0.0, 2)

        # Buffer confidence
        buf_scores = [b.get('score', 0) for b in results.get('buffer', {}).get('top_buffers', []) if isinstance(b.get('score', None), (int, float))]
        scores['buffer'] = round(min(1.0, max(0.0, buf_scores[0] / 100)) if buf_scores else 0.0, 2)

        # Solvent confidence
        sol_scores = [s.get('score', 0) for s in results.get('solvent', {}).get('top_solvents', []) if isinstance(s.get('score', None), (int, float))]
        scores['solvent'] = round(min(1.0, max(0.0, sol_scores[0] / 100)) if sol_scores else 0.0, 2)

        # Column confidence
        col_scores = [c.get('score', 0) for c in results.get('column', {}).get('top_columns', []) if isinstance(c.get('score', None), (int, float))]
        scores['column'] = round(min(1.0, max(0.0, col_scores[0] / 100)) if col_scores else 0.0, 2)

        # Overall confidence (mean of available scores)
        score_values = [v for v in scores.values() if isinstance(v, (int, float))]
        scores['overall'] = round((sum(score_values) / len(score_values)) if score_values else 0.0, 2)

        results['scores'] = scores

        # Build detailed rationale text for each component (for report interpretation)
        rationale = {
            'physchem': (
                f"Physicochemical predictions use RDKit descriptors (LogP, TPSA, H-bonding) "
                f"and rule-based filters (Lipinski violations={lipinski_violations}). "
                f"Confidence is based on QED score ({qed:.3f})."
            ) if phys else '',
            'hsm': (
                "HSM descriptors (eta', sigma', beta', alpha', kappa') are calculated to "
                "capture hydrophobicity, steric resistance, and hydrogen bonding interactions."
            ) if hsm else '',
            'buffer': (
                "Buffer recommendations are based on pKa matching, buffering range, and "
                "compatibility with common HPLC additives (e.g., formic acid, ammonium acetate)."
            ) if buf_scores else '',
            'solvent': (
                "Solvent scoring considers polarity, hydrogen-bonding character, and "
                "miscibility to prioritize common HPLC solvents (MeOH, ACN, EtOH)."
            ) if sol_scores else '',
            'column': (
                "Column ranking uses HSM-derived retention parameters and known selectivity "
                "profiles to recommend suitable reversed-phase stationary phases."
            ) if col_scores else ''
        }
        results['rationale'] = rationale

        # Attach module references for the report
        results['references'] = self.reference_manager.get_all_references()
        results['references_by_module'] = {
            module: self.reference_manager.get_references_by_module(module)
            for module in self.reference_manager.module_refs.keys()
        }

    def _add_to_history(self, results: Dict):
        """Add results to history."""
        history_entry = {
            'TIMESTAMP': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'COMPOUND NAME': results.get('name', ''),
            'SMILES': results.get('smiles', '')[:50] + '...' if len(results.get('smiles', '')) > 50 else results.get('smiles', ''),
            'PROJECT': results.get('project', ''),
            'STATUS': results.get('status', ''),
            'LOGP': results.get('physchem', {}).get('logp', 'N/A')
        }
        
        self.history.insert(0, history_entry)
        self.history = self.history[:50]  # Keep last 50 entries
    
    def process_batch_queue(self) -> List[Dict]:
        """
        Process all compounds in the batch queue.
        
        Returns:
            List of results for each compound
        """
        batch_queue = self.read_batch_queue()
        
        if not batch_queue:
            logger.warning("Batch queue is empty")
            return []
        
        logger.info(f"Processing batch queue with {len(batch_queue)} compounds")
        
        results_list = []
        
        for i, entry in enumerate(batch_queue):
            logger.info(f"Processing batch item {i+1}/{len(batch_queue)}: {entry['name'] or entry['smiles'][:30]}")
            
            # Update status in Excel
            self._update_batch_status(i+1, 'Processing', f"{int((i/len(batch_queue))*100)}%")
            
            try:
                # Process compound
                result = self.process_single_compound(
                    smiles=entry['smiles'],
                    name=entry['name'],
                    project="Batch Processing",
                    notes=""
                )
                
                results_list.append(result)
                
                # Update status
                status = 'Completed' if result['status'] == 'Success' else 'Failed'
                self._update_batch_status(i+1, status, '100%')
                
            except Exception as e:
                logger.error(f"Error processing batch item {i+1}: {e}")
                self._update_batch_status(i+1, 'Failed', '0%')
                results_list.append({'error': str(e), 'status': 'Failed'})
        
        # Update final progress
        self._update_progress('100%', 'Complete')
        
        return results_list
    
    def _update_batch_status(self, row_offset: int, status: str, progress: str):
        """Update status and progress for a batch queue item."""
        if not self.excel_file:
            return
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)
            sheet_name = self._get_existing_sheet_name(['DASHBOARD', 'Input'], 'Input')
            ws = wb[sheet_name]
            
            # Batch queue starts at row 14
            status_row = 13 + row_offset
            progress_row = 13 + row_offset
            
            ws.cell(row=status_row, column=4, value=status)  # Column D = Status
            ws.cell(row=progress_row, column=5, value=progress)  # Column E = Progress
            
            wb.save(self.excel_file)
            
        except Exception as e:
            logger.error(f"Error updating batch status: {e}")
    
    def _update_progress(self, progress: str, eta: str):
        """Update progress bar and ETA in Excel."""
        if not self.excel_file:
            return
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)
            sheet_name = self._get_existing_sheet_name(['DASHBOARD', 'Input'], 'DASHBOARD')
            ws = wb[sheet_name]

            ws['H8'] = "Progress: " + str(progress)
            ws['H7'] = "ETA: " + str(eta)
            
            wb.save(self.excel_file)
            
        except Exception as e:
            logger.error(f"Error updating progress: {e}")
    
    def generate_report(self, results: Dict, output_file: Optional[str] = None) -> str:
        """
        Generate comprehensive report with all results, references, and professional formatting.
        
        Uses ReportTemplate and ExcelStyleManager for consistent styling across sheets.
        
        Args:
            results: Results dictionary from process_single_compound
            output_file: Optional output file path
            
        Returns:
            Path to generated report
        """
        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            name_part = results.get('name', 'compound').replace(' ', '_')
            output_file = self.reports_dir / f"report__{timestamp}.xlsx"
        else:
            # Normalize output path: allow directory or base name without extension
            output_file = Path(output_file)
            if (str(output_file).endswith("/") or str(output_file).endswith("\\")) or (output_file.exists() and output_file.is_dir()):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = output_file / f"report__{timestamp}.xlsx"
            elif output_file.suffix == "":
                output_file = output_file.with_suffix(".xlsx")
            output_file.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Generating formatted report: {output_file}")
        
        try:
            # Create workbook with openpyxl for formatting control
            from openpyxl import Workbook
            wb = Workbook()
            
            # Initialize template
            template = ReportTemplate() if HAS_REPORT_TEMPLATE else None
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets in order
            # Sheet 1: Cover
            if template:
                ws = wb.create_sheet('Cover', 0)
                template.create_cover_sheet(ws, results)
            
            # Sheet 2: Summary
            if template:
                ws = wb.create_sheet('Summary', 1)
                template.create_summary_sheet(ws, results)
            
            # Sheet 3: Physicochemical Properties
            if 'physchem' in results:
                if template:
                    ws = wb.create_sheet('Physicochemical', 2)
                    template.create_physchem_sheet(ws, results)
                else:
                    self._create_physchem_sheet_legacy(wb, results)
            
            # Sheet 4: HSM Descriptors
            if 'hsm' in results:
                if template:
                    ws = wb.create_sheet('HSM', 3)
                    template.create_hsm_sheet(ws, results)
                else:
                    self._create_hsm_sheet_legacy(wb, results)
            
            # Sheet 5: Column Recommendations
            if 'column' in results:
                if template:
                    ws = wb.create_sheet('Columns', 4)
                    template.create_column_sheet(ws, results)
                else:
                    self._create_column_sheet_legacy(wb, results)
            
            # Sheet 6: Solvent Recommendations
            if 'solvent' in results:
                ws = wb.create_sheet('Solvents', 5)
                if template and hasattr(template, 'create_solvent_sheet'):
                    template.create_solvent_sheet(ws, results)
                else:
                    self._create_solvent_sheet_sheet(ws, results)
            
            # Sheet 7: Buffer Recommendations
            if 'buffer' in results:
                ws = wb.create_sheet('Buffers', 6)
                if template and hasattr(template, 'create_buffer_sheet'):
                    template.create_buffer_sheet(ws, results)
                else:
                    self._create_buffer_sheet_sheet(ws, results)
            
            # Sheet 8: Gradient Program
            if 'gradient' in results:
                ws = wb.create_sheet('Gradient', 7)
                if template and hasattr(template, 'create_gradient_sheet'):
                    template.create_gradient_sheet(ws, results)
                else:
                    self._create_gradient_sheet_sheet(ws, results)
            
            # Sheet 9: References
            if self.settings.get('include_references', True):
                ws = wb.create_sheet('References', 8)
                self._create_references_sheet_sheet(ws)
            
            # Sheet 10: Metadata
            ws = wb.create_sheet('Metadata', 9)
            self._create_method_parameters_sheet_sheet(ws)
            
            # Save workbook
            wb.save(output_file)
            logger.info(f"Report generated: {output_file}")
            return str(output_file)
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            traceback.print_exc()
            
            # Fallback to JSON
            json_file = Path(output_file).with_suffix('.json')
            with open(json_file, 'w') as f:
                json.dump(results, f, indent=2, default=str)
            logger.info(f"JSON fallback generated: {json_file}")
            return str(json_file)
    
    # ========================================================================
    # SHEET CREATION METHODS (for Worksheet objects)
    # ========================================================================
    
    def _create_physchem_sheet_legacy(self, wb, results: Dict):
        """Create physicochemical properties sheet in workbook."""
        ws = wb.create_sheet('Physicochemical', 2)
        physchem = results.get('physchem', {})
        
        # Write headers
        ws['A1'] = 'Property'
        ws['B1'] = 'Value'
        
        row = 2
        for key, value in physchem.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
    
    def _create_hsm_sheet_legacy(self, wb, results: Dict):
        """Create HSM descriptors sheet in workbook."""
        ws = wb.create_sheet('HSM', 3)
        hsm = results.get('hsm', {})
        
        # Create data
        data = [
            ('eta\' (Hydrophobicity)', hsm.get('eta_prime', 'N/A')),
            ('sigma\' (Steric Resistance)', hsm.get('sigma_prime', 'N/A')),
            ('beta\' (H-Bond Basicity)', hsm.get('beta_prime', 'N/A')),
            ('alpha\' (H-Bond Acidity)', hsm.get('alpha_prime', 'N/A')),
            ('kappa\' (Cationic Charge)', hsm.get('kappa_prime', 'N/A')),
        ]
        
        # Write headers
        ws['A1'] = 'Descriptor'
        ws['B1'] = 'Value'
        
        row = 2
        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
    
    def _create_column_sheet_legacy(self, wb, results: Dict):
        """Create column recommendations sheet in workbook."""
        ws = wb.create_sheet('Columns', 4)
        column = results.get('column', {})
        top_columns = column.get('top_columns', [])
        
        # Write headers
        headers = ['Rank', 'Column Name', 'Manufacturer', 'H', 'S', 'A', 'B', 'C', 'Score']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, col in enumerate(top_columns, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Rank
            ws.cell(row=row_idx, column=2, value=col.get('name', ''))
            ws.cell(row=row_idx, column=3, value=col.get('manufacturer', ''))
            ws.cell(row=row_idx, column=4, value=col.get('h', 0))
            ws.cell(row=row_idx, column=5, value=col.get('s', 0))
            ws.cell(row=row_idx, column=6, value=col.get('a', 0))
            ws.cell(row=row_idx, column=7, value=col.get('b', 0))
            ws.cell(row=row_idx, column=8, value=col.get('c', 0))
            ws.cell(row=row_idx, column=9, value=col.get('score', 0))
    
    def _create_solvent_sheet_sheet(self, ws, results: Dict):
        """Create solvent recommendations sheet."""
        solvent = results.get('solvent', {})
        top_solvents = solvent.get('top_solvents', [])
        
        # Write headers
        ws['A1'] = 'Rank'
        ws['B1'] = 'Solvent Name'
        ws['C1'] = 'Properties'
        
        row = 2
        for i, solv in enumerate(top_solvents, 1):
            ws[f'A{row}'] = i
            ws[f'B{row}'] = solv.get('name', '')
            ws[f'C{row}'] = str(solv.get('properties', {}))
            row += 1
    
    def _create_buffer_sheet_sheet(self, ws, results: Dict):
        """Create buffer recommendations sheet."""
        buffer_data = results.get('buffer', {})
        top_buffers = buffer_data.get('top_buffers', [])
        
        # Write headers
        ws['A1'] = 'Rank'
        ws['B1'] = 'Buffer Name'
        ws['C1'] = 'pH'
        
        row = 2
        for i, buf in enumerate(top_buffers, 1):
            ws[f'A{row}'] = i
            ws[f'B{row}'] = buf.get('name', '')
            ws[f'C{row}'] = buf.get('pH', '')
            row += 1
    
    def _create_gradient_sheet_sheet(self, ws, results: Dict):
        """Create gradient program sheet."""
        gradient = results.get('gradient', {})
        
        # Write basic info
        ws['A1'] = 'Gradient Parameter'
        ws['B1'] = 'Value'
        ws['A2'] = 'Total Runtime'
        ws['B2'] = gradient.get('total_runtime', 'N/A')
        ws['A3'] = 'Initial Composition'
        ws['B3'] = gradient.get('initial_composition', 'N/A')
        ws['A4'] = 'Final Composition'
        ws['B4'] = gradient.get('final_composition', 'N/A')
    
    def _create_references_sheet_sheet(self, ws):
        """Create references sheet."""
        references = self.reference_manager.get_all_references()

        # Write headers
        ws['A1'] = 'Reference (Vancouver)'
        ws.column_dimensions['A'].width = 120

        row = 2
        for idx, ref in enumerate(references[:200], 1):  # Allow more entries but limit
            try:
                # If Reference object
                if hasattr(ref, 'format_citation'):
                    text = f"{idx}. {ref.format_citation('vancouver')}"
                elif isinstance(ref, dict):
                    authors = ref.get('authors', [])
                    if isinstance(authors, list):
                        authors = ', '.join(authors[:3])
                    text = f"{idx}. {authors}. {ref.get('title','')}. {ref.get('journal','')}. {ref.get('year','')}"
                    if ref.get('doi'):
                        text += f" doi:{ref.get('doi')}"
                else:
                    text = f"{idx}. {str(ref)}"
            except Exception:
                text = f"{idx}. {str(ref)}"

            ws[f'A{row}'] = text
            row += 1
    
    def _create_method_parameters_sheet_sheet(self, ws):
        """Create method parameters metadata sheet."""
        ws['A1'] = 'Parameter'
        ws['B1'] = 'Value'
        
        # Generate session ID if not exists
        session_id = getattr(self, 'session_id', datetime.now().strftime('%Y%m%d_%H%M%S'))
        
        metadata = [
            ('CHROME-pred Version', '2.0'),
            ('Analysis Date', str(datetime.now())),
            ('Report ID', session_id),
        ]
        
        row = 2
        for param, value in metadata:
            ws[f'A{row}'] = param
            ws[f'B{row}'] = value
            row += 1
    
    # ========================================================================
    # LEGACY SHEET CREATION METHODS (keep for compatibility)
    # ========================================================================
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create summary sheet."""
        summary_data = [{
            'Property': 'Compound Name',
            'Value': results.get('name', 'N/A')
        }, {
            'Property': 'SMILES',
            'Value': results.get('smiles', 'N/A')
        }, {
            'Property': 'Project',
            'Value': results.get('project', 'N/A')
        }, {
            'Property': 'Status',
            'Value': results.get('status', 'N/A')
        }, {
            'Property': 'Timestamp',
            'Value': results.get('timestamp', 'N/A')
        }, {
            'Property': 'Molecular Formula',
            'Value': results.get('physchem', {}).get('molecular_formula', 'N/A')
        }, {
            'Property': 'Molecular Weight',
            'Value': results.get('physchem', {}).get('molecular_weight', 'N/A')
        }, {
            'Property': 'LogP',
            'Value': results.get('physchem', {}).get('logp', 'N/A')
        }, {
            'Property': 'LogD (pH 7.4)',
            'Value': results.get('physchem', {}).get('logd_ph74', 'N/A')
        }, {
            'Property': 'TPSA',
            'Value': results.get('physchem', {}).get('tpsa', 'N/A')
        }, {
            'Property': 'Lipinski Violations',
            'Value': results.get('physchem', {}).get('lipinski_violations', 'N/A')
        }, {
            'Property': 'Top Column',
            'Value': results.get('column', {}).get('top_columns', [{}])[0].get('name', 'N/A') if results.get('column', {}).get('top_columns') else 'N/A'
        }, {
            'Property': 'Top Solvent',
            'Value': results.get('solvent', {}).get('top_solvents', [{}])[0].get('name', 'N/A') if results.get('solvent', {}).get('top_solvents') else 'N/A'
        }, {
            'Property': 'Top Buffer',
            'Value': results.get('buffer', {}).get('top_buffers', [{}])[0].get('name', 'N/A') if results.get('buffer', {}).get('top_buffers') else 'N/A'
        }, {
            'Property': 'Total Runtime',
            'Value': f"{results.get('gradient', {}).get('total_runtime', 'N/A')} min"
        }]
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_physchem_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create physicochemical properties sheet."""
        physchem = results.get('physchem', {})
        
        data = []
        for key, value in physchem.items():
            data.append({'Property': key.replace('_', ' ').title(), 'Value': value})
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='Physicochemical', index=False)
        
        # Add chromatographic descriptors if available
        if 'chrom_descriptors' in results:
            chrom_data = []
            for key, value in results['chrom_descriptors'].items():
                chrom_data.append({'Property': key.replace('_', ' ').title(), 'Value': value})
            
            chrom_df = pd.DataFrame(chrom_data)
            chrom_df.to_excel(writer, sheet_name='Physicochemical', 
                             startrow=len(df)+3, index=False)
    
    def _create_hsm_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create HSM descriptors sheet."""
        hsm = results.get('hsm', {})
        
        data = [{
            'Descriptor': 'η′ (Hydrophobicity)',
            'Value': hsm.get('eta_prime', 'N/A'),
            'Description': 'Hydrophobicity parameter'
        }, {
            'Descriptor': 'σ′ (Steric Resistance)',
            'Value': hsm.get('sigma_prime', 'N/A'),
            'Description': 'Steric/bulkiness parameter'
        }, {
            'Descriptor': 'β′ (H-Bond Basicity)',
            'Value': hsm.get('beta_prime', 'N/A'),
            'Description': 'Hydrogen bond basicity'
        }, {
            'Descriptor': 'α′ (H-Bond Acidity)',
            'Value': hsm.get('alpha_prime', 'N/A'),
            'Description': 'Hydrogen bond acidity'
        }, {
            'Descriptor': 'κ′ (Cationic Charge)',
            'Value': hsm.get('kappa_prime', 'N/A'),
            'Description': 'Cationic charge at pH 7.0'
        }]
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name='HSM Descriptors', index=False)
    
    def _create_column_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create column recommendations sheet."""
        column = results.get('column', {})
        top_columns = column.get('top_columns', [])
        
        if top_columns:
            data = []
            for i, col in enumerate(top_columns, 1):
                data.append({
                    'Rank': i,
                    'Column Name': col.get('name', ''),
                    'Manufacturer': col.get('manufacturer', ''),
                    'H': col.get('h', 0),
                    'S': col.get('s', 0),
                    'A': col.get('a', 0),
                    'B': col.get('b', 0),
                    'C': col.get('c', 0),
                    'Score': col.get('score', 0)
                })
            
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='Column Recommendations', index=False)
            
            # Add advice
            advice = column.get('advice', [])
            if advice:
                advice_df = pd.DataFrame([{'Advice': a} for a in advice])
                advice_df.to_excel(writer, sheet_name='Column Recommendations', 
                                  startrow=len(df)+3, index=False)
    
    def _create_solvent_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create solvent recommendations sheet."""
        solvent = results.get('solvent', {})
        top_solvents = solvent.get('top_solvents', [])
        
        if top_solvents:
            data = []
            for i, sol in enumerate(top_solvents, 1):
                data.append({
                    'Rank': i,
                    'Solvent': sol.get('name', ''),
                    'Score': sol.get('score', 0)
                })
            
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='Solvent Recommendations', index=False)
            
            # Add initial composition
            comp = solvent.get('initial_composition', {})
            if comp:
                comp_data = [{
                    'Parameter': 'Initial Organic %',
                    'Value': comp.get('organic_percent', 'N/A'),
                    'Rationale': comp.get('rationale', '')
                }]
                comp_df = pd.DataFrame(comp_data)
                comp_df.to_excel(writer, sheet_name='Solvent Recommendations', 
                                startrow=len(df)+3, index=False)
    
    def _create_buffer_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create buffer recommendations sheet."""
        buffer = results.get('buffer', {})
        top_buffers = buffer.get('top_buffers', [])
        
        if top_buffers:
            data = []
            for i, buf in enumerate(top_buffers, 1):
                data.append({
                    'Rank': i,
                    'Buffer': buf.get('name', ''),
                    'pKa': buf.get('pka', 'N/A'),
                    'Score': buf.get('score', 0)
                })
            
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name='Buffer Recommendations', index=False)
    
    def _create_gradient_sheet(self, writer: pd.ExcelWriter, results: Dict):
        """Create gradient program sheet."""
        gradient = results.get('gradient', {})
        
        summary_data = [{
            'Parameter': 'Total Runtime',
            'Value': f"{gradient.get('total_runtime', 'N/A')} min"
        }, {
            'Parameter': 'Peak Capacity',
            'Value': gradient.get('peak_capacity', 'N/A')
        }, {
            'Parameter': 'Minimum Resolution',
            'Value': gradient.get('min_resolution', 'N/A')
        }]
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Gradient Program', index=False)
        
        # Gradient segments
        segments = gradient.get('segments', [])
        if segments:
            seg_data = []
            for i, seg in enumerate(segments, 1):
                seg_data.append({
                    'Segment': i,
                    'Start Time (min)': seg.get('start_time', 0),
                    'End Time (min)': seg.get('end_time', 0),
                    'Start %B': seg.get('start_b', 0),
                    'End %B': seg.get('end_b', 0),
                    'Duration (min)': seg.get('end_time', 0) - seg.get('start_time', 0)
                })
            
            seg_df = pd.DataFrame(seg_data)
            seg_df.to_excel(writer, sheet_name='Gradient Program', 
                           startrow=len(summary_df)+3, index=False)
    
    def _create_references_sheet(self, writer: pd.ExcelWriter):
        """Create references sheet with all citations."""
        references = self.reference_manager.get_all_references()
        
        if references:
            ref_data = []
            for i, ref in enumerate(references, 1):
                if 'authors' in ref:
                    citation = f"{ref.get('authors', '')} ({ref.get('year', 'n.d.')}). {ref.get('title', '')}. {ref.get('journal', '')}, {ref.get('volume', '')}, {ref.get('pages', '')}."
                elif 'database' in ref:
                    citation = f"{ref.get('database', '')}. Available at: {ref.get('url', '')}"
                else:
                    citation = ref.get('citation', '')
                
                ref_data.append({
                    '#': i,
                    'Module': ref.get('module', ''),
                    'Citation': citation,
                    'DOI': ref.get('doi', '')
                })
            
            df = pd.DataFrame(ref_data)
            df.to_excel(writer, sheet_name='References', index=False)
    
    def _create_method_parameters_sheet(self, writer: pd.ExcelWriter):
        """Create method parameters sheet."""
        params_data = [{
            'Parameter': 'Column Length',
            'Value': '150 mm',
            'Description': 'Standard analytical column'
        }, {
            'Parameter': 'Column ID',
            'Value': '4.6 mm',
            'Description': 'Standard internal diameter'
        }, {
            'Parameter': 'Particle Size',
            'Value': '3.5 µm',
            'Description': 'Typical for U/HPLC'
        }, {
            'Parameter': 'Flow Rate',
            'Value': '1.5 mL/min',
            'Description': 'Default flow rate'
        }, {
            'Parameter': 'Temperature',
            'Value': '30°C',
            'Description': 'Default column temperature'
        }, {
            'Parameter': 'Detection Wavelength',
            'Value': '254 nm',
            'Description': 'Default UV detection'
        }, {
            'Parameter': 'Injection Volume',
            'Value': '5 µL',
            'Description': 'Standard injection'
        }, {
            'Parameter': 'Buffer Concentration',
            'Value': '50 mM',
            'Description': 'Default buffer strength'
        }]
        
        df = pd.DataFrame(params_data)
        df.to_excel(writer, sheet_name='Method Parameters', index=False)
    
    def update_excel_results(self, results: Dict, report_paths: Optional[Dict[str, str]] = None):
        """Update Excel file with results for the current compound.

        Args:
            results: Results dictionary from process_single_compound
            report_paths: Optional dict with keys 'excel' and 'pdf' containing report paths
        """
        if not self.excel_file:
            return

        try:
            wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)

            # Ensure master workbook structure (sheets and headers)
            self._ensure_master_workbook_structure(wb)

            # Update Dashboard/Input sheet status
            sheet_name = self._get_existing_sheet_name(['DASHBOARD', 'Input'], 'Input')
            ws_input = wb[sheet_name]
            ws_input['H2'] = "✅ Completed"
            ws_input['H4'] = f"Last Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

            # Append to Search History sheet (audit trail)
            ws_history = self._ensure_search_history_sheet(wb)
            last_row = ws_history.max_row + 1

            report_id = results.get('report_id', '')
            overall_conf = results.get('scores', {}).get('overall', '')
            phys_conf = results.get('scores', {}).get('physchem', '')
            hsm_conf = results.get('scores', {}).get('hsm', '')
            buffer_conf = results.get('scores', {}).get('buffer', '')
            solvent_conf = results.get('scores', {}).get('solvent', '')
            column_conf = results.get('scores', {}).get('column', '')

            history_entry = [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                report_id,
                results.get('smiles', ''),
                results.get('name', ''),
                results.get('project', ''),
                results.get('status', ''),
                overall_conf,
                phys_conf,
                hsm_conf,
                buffer_conf,
                solvent_conf,
                column_conf,
                results.get('processing_time_s', ''),
                '',  # Excel report hyperlink placeholder
                '',  # PDF report hyperlink placeholder
                json.dumps(results, default=str)  # Raw data
            ]

            for col, value in enumerate(history_entry, 1):
                cell = ws_history.cell(row=last_row, column=col, value=value)

            # Hyperlinks
            if report_paths:
                excel_path = report_paths.get('excel')
                pdf_path = report_paths.get('pdf')
                if excel_path:
                    cell = ws_history.cell(row=last_row, column=14)
                    cell.value = 'Open'
                    cell.hyperlink = str(excel_path)
                if pdf_path:
                    cell = ws_history.cell(row=last_row, column=15)
                    cell.value = 'Open'
                    cell.hyperlink = str(pdf_path)

            # Hide raw data column
            ws_history.column_dimensions['P'].hidden = True

            wb.save(self.excel_file)
            logger.info("Excel file updated with results")

        except Exception as e:
            logger.error(f"Error updating Excel: {e}")
    
    def run_interactive(self, excel_file: Union[str, Path]):
        """
        Run interactive session with Excel frontend.
        
        Args:
            excel_file: Path to Method_Development_Input.xlsm
        """
        # Load Excel file
        if not self.load_excel_file(excel_file):
            logger.error("Failed to load Excel file")
            return False
        
        logger.info("=" * 80)
        logger.info("INTERACTIVE MODE - READY")
        logger.info("=" * 80)
        logger.info("Waiting for Excel commands...")
        
        # Main loop - this would be triggered by Excel VBA
        # For now, we'll just process a single compound as example
        
        # Read input
        smiles, name, project, notes = self.read_input_smiles()
        
        if smiles:
            # Validate SMILES
            is_valid, message = self.validate_smiles(smiles)
            self._update_progress('50%', 'Processing')
            
            if is_valid:
                # Update live preview
                self.update_live_preview(smiles)
                
                # Process compound
                results = self.process_single_compound(smiles, name, project, notes)
                
                # Generate reports (Excel + PDF)
                report_paths = self.generate_reports(results, formats=['excel', 'pdf'])
                
                # Update Excel (history & dashboard)
                self.update_excel_results(results, report_paths)
                
                self._update_progress('100%', 'Complete')
                
                logger.info(f"Report generated: {report_paths}")
                
                return True
            else:
                logger.error(f"Invalid SMILES: {message}")
                wb = openpyxl.load_workbook(self.excel_file, keep_vba=True)
                ws = wb['Input']
                ws['H3'] = f"❌ Invalid: {message[:30]}"
                wb.save(self.excel_file)
                return False

        return True

    def generate_pdf_report(self, results: Dict, output_file: Optional[str] = None) -> str:
        """
        Generate comprehensive PDF report with all results and professional formatting.
        
        Requires reportlab library. If not available, returns a text fallback.
        
        Args:
            results: Results dictionary from process_single_compound
            output_file: Optional output file path
        
        Returns:
            Path to generated PDF file, or text fallback path if PDF generation is unavailable.
        """
        pdf_generator = None
        if HAS_REPORT_TEMPLATE:
            try:
                pdf_generator = PDFReportGenerator()
            except Exception as e:
                logger.warning(f"PDFReportGenerator import failed: {e}")
                pdf_generator = None

        if pdf_generator:
            try:
                if not getattr(pdf_generator, 'available', True):
                    logger.warning("PDF generator unavailable; falling back to built-in fallback generator")
                    pdf_generator = None
            except Exception:
                pdf_generator = None

        if output_file is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = self.reports_dir / f"report__{timestamp}.pdf"
        else:
            output_file = Path(output_file)
            if (str(output_file).endswith("/") or str(output_file).endswith("\\")) or (output_file.exists() and output_file.is_dir()):
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = output_file / f"report__{timestamp}.pdf"
            elif output_file.suffix == "":
                output_file = output_file.with_suffix(".pdf")
            output_file.parent.mkdir(parents=True, exist_ok=True)

        references = self.reference_manager.get_all_references() if self.reference_manager else []

        if pdf_generator:
            try:
                pdf_path = pdf_generator.generate_pdf(results, references, str(output_file))
                if pdf_path:
                    logger.info(f"PDF report generated: {pdf_path}")
                    return pdf_path
            except Exception as e:
                logger.warning(f"PDF generation with template generator failed: {e}")

        if HAS_REPORTLAB:
            pdf_fallback = SimplePDFReportGenerator()
            pdf_path = pdf_fallback.generate_pdf(results, references, str(output_file))
            if pdf_path:
                logger.info(f"PDF report generated with fallback: {pdf_path}")
                return pdf_path
            logger.warning("Fallback PDF generation failed")

        logger.warning("PDF generation unavailable, creating text fallback")
        txt_path = Path(str(output_file)).with_suffix('.txt')
        try:
            txt_path.parent.mkdir(parents=True, exist_ok=True)
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write('CHROME-pred Report (text fallback)\n')
                f.write('Generated: ' + datetime.now().isoformat() + '\n\n')
                f.write(json.dumps(results, indent=2, default=str))
            logger.info(f"Text fallback report generated: {txt_path}")
            return str(txt_path)
        except Exception as e:
            logger.error(f"Error writing text fallback report: {e}")
            return ""

    def generate_reports(self, results: Dict, formats: List[str] = None, output_base: Optional[str] = None) -> Dict[str, str]:
        """
        Generate reports in multiple formats (Excel and/or PDF).
        Ensures both files use the same base name if --format both is used.
        """
        if formats is None:
            formats = ['excel']
        if isinstance(formats, str):
            if formats.lower() == 'both':
                formats = ['excel', 'pdf']
            else:
                formats = [formats]
        formats = [f.lower() for f in formats]

        # Determine base path for both outputs
        base_path = None
        if output_base:
            base_path = Path(output_base)
            if base_path.suffix.lower() in ['.xlsx', '.xlsm', '.xls', '.pdf']:
                base_path = base_path.with_suffix('')
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            name_part = results.get('name', 'compound').replace(' ', '_')
            base_path = Path(self.reports_dir) / f"report__{timestamp}"

        report_paths = {}
        # Always generate Excel if requested
        if 'excel' in formats:
            excel_path = self.generate_report(results, str(base_path.with_suffix('.xlsx')))
            if excel_path:
                report_paths['excel'] = str(excel_path)
                logger.info(f"Excel report: {excel_path}")
        # Always generate PDF if requested
        if 'pdf' in formats:
            pdf_path = self.generate_pdf_report(results, str(base_path.with_suffix('.pdf')))
            if pdf_path:
                report_paths['pdf'] = str(pdf_path)
                logger.info(f"PDF report: {pdf_path}")
        return report_paths


def main():

    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Chromatography AI System v2.0')
    parser.add_argument('--excel', '-e', type=str, 
                       default='Method_Development_Input.xlsm',
                       help='Path to Excel input file')
    parser.add_argument('--smiles', '-s', type=str,
                       help='Process single SMILES string')
    parser.add_argument('--project', '-p', type=str, default='',
                       help='Project name')
    parser.add_argument('--smiles-file', type=str,
                       help='Path to a text file containing a single SMILES string')
    parser.add_argument('--batch', '-b', action='store_true',
                       help='Process batch queue from Excel')
    parser.add_argument('--output', '-o', type=str,
                       help='Output report file')
    parser.add_argument('--validate-only', action='store_true',
                       help='Validate SMILES only, no full report')
    parser.add_argument('--format', '-f', type=str, default='excel',
                       choices=['excel', 'pdf', 'both'],
                       help='Report format: excel, pdf, or both (default: excel)')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Verbose output')
    
    args = parser.parse_args()
    
    # Initialize controller
    controller = ChromatographyAIController()

    # FIX PY-002: handle --validate-only (SMILES validation only, no report)
    if getattr(args, 'validate_only', False):
        smiles_to_check = None
        if args.smiles_file:
            try:
                with open(args.smiles_file, 'r') as f:
                    smiles_to_check = f.read().strip()
            except Exception:
                print("ERROR: Could not read SMILES file")
                return 1
        elif args.smiles:
            smiles_to_check = args.smiles
        if smiles_to_check:
            is_valid, msg = controller.validate_smiles(smiles_to_check)
            print("OK|VALID|" + smiles_to_check if is_valid else "ERROR|INVALID|" + msg)
        else:
            print("ERROR: No SMILES provided")
        return 0
    
    # Support SMILES passed directly or via a temporary file (VBA compatibility)
    smiles_input = None
    if args.smiles_file:
        try:
            with open(args.smiles_file, 'r') as f:
                smiles_input = f.read().strip()
        except Exception as e:
            logger.error(f"Error reading SMILES file {args.smiles_file}: {e}")
            print(f"Error reading SMILES file: {e}")
            return 1
    elif args.smiles:
        smiles_input = args.smiles

    if smiles_input:
        # Process single SMILES from command line or file
        results = controller.process_single_compound(smiles_input, project=getattr(args, 'project', ''))
        
        # If the user supplied an output file with a PDF extension, assume PDF format
        desired_format = args.format.lower()
        if desired_format == 'excel' and args.output:
            out_lc = args.output.lower()
            if out_lc.endswith('.pdf'):
                desired_format = 'pdf'
            elif out_lc.endswith(('.xlsx', '.xlsm', '.xls')):
                desired_format = 'excel'

        # Generate reports in requested formats
        report_paths = {}
        if desired_format == 'excel':
            report_paths['excel'] = controller.generate_report(results, args.output)
        elif desired_format == 'pdf':
            report_paths['pdf'] = controller.generate_pdf_report(results, args.output)
        else:  # both
            report_paths = controller.generate_reports(results, ['excel', 'pdf'], args.output)

        # Print a machine-readable result line for VBA parsing
        scores = results.get('scores', {})
        result_line = "|".join([
            "OK",
            results.get('report_id', ''),
            results.get('smiles', ''),
            results.get('name', ''),
            results.get('project', ''),
            str(scores.get('overall', '')),
            str(scores.get('physchem', '')),
            str(scores.get('hsm', '')),
            str(scores.get('buffer', '')),
            str(scores.get('solvent', '')),
            str(scores.get('column', '')),
            str(results.get('processing_time_s', '')),
            report_paths.get('excel', ''),
            report_paths.get('pdf', ''),
        ])
        print(result_line)
        
    elif args.batch:
        # Process batch queue
        if not controller.load_excel_file(args.excel):
            print("Failed to load Excel file")
            return 1
        
        results_list = controller.process_batch_queue()
        print(f"Processed {len(results_list)} compounds")
        
    else:
        # Interactive mode with Excel
        if not controller.run_interactive(args.excel):
            print("Interactive mode failed")
            return 1
    
    return 0



if __name__ == "__main__":
    sys.exit(main())